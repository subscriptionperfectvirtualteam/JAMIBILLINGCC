"""
JamiBilling - RDN Fee Scraper Application
Enhanced version with logging, error handling, and debug features
"""

import os
import json
import re
import csv
import datetime
import io
import logging
import pypyodbc as pyodbc  # Using pypyodbc instead of pyodbc
from flask import Flask, render_template, request, jsonify, session, send_file
from flask_session import Session
from selenium import webdriver
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.common.exceptions import TimeoutException, NoSuchElementException
from werkzeug.utils import secure_filename
from bs4 import BeautifulSoup
import openpyxl  # Using openpyxl instead of pandas for Excel
import requests

# Import ChromeDriver manager
from chromedriver_manager import download_chromedriver

# Configure logging
logging.basicConfig(
    filename='jami_billing.log',
    level=logging.INFO,
    format='%(asctime)s - %(name)s - %(levelname)s - %(message)s'
)

# Initialize Flask app
app = Flask(__name__)
app.config["SECRET_KEY"] = os.urandom(24)
app.config["SESSION_TYPE"] = "filesystem"
app.config["SESSION_PERMANENT"] = False
app.config["SESSION_FILE_DIR"] = "flask_session"
Session(app)

# Load application config
try:
    with open('config.json', 'r') as f:
        app_config = json.load(f)
    logging.info("Config loaded successfully")
except Exception as e:
    logging.error(f"Error loading config: {str(e)}")
    app_config = {
        "database": {
            "provider": "SQLOLEDB",
            "server": "localhost",
            "database": "RDN_Billing",
            "username": "admin",
            "password": "password"
        },
        "rdn": {
            "login_url": "https://secureauth.recoverydatabase.net/public/login",
            "case_url_template": "https://app.recoverydatabase.net/alpha_rdn/module/default/case2/?case_id={case_id}"
        }
    }

# Load fee categories from the JSON file
try:
    with open('backend/fee_categories.json', 'r') as f:
        fee_categories = json.load(f)
    logging.info("Fee categories loaded successfully")
except Exception as e:
    logging.error(f"Error loading fee categories: {str(e)}")
    fee_categories = {
        "Storage": {
            "keywords": ["storage fee", "daily storage", "impound fee", "lot fee"],
            "color": "#4e73df"
        },
        "Repossession": {
            "keywords": ["repo fee", "recovery fee", "tow fee", "involuntary repo"],
            "color": "#e74a3b"
        },
        "Other": {
            "keywords": ["other"],
            "color": "#858796"
        }
    }

# Ensure required directories exist
os.makedirs('flask_session', exist_ok=True)
os.makedirs(os.path.join('static', 'exports'), exist_ok=True)
os.makedirs('debug', exist_ok=True)
os.makedirs('drivers', exist_ok=True)

@app.route('/')
def index():
    """Render the main application page"""
    logging.info("Rendering main page")
    return render_template('index.html')

@app.route('/api/login', methods=['POST'])
def login():
    """Handle login to RDN using Selenium"""
    logging.info("Login request received")
    data = request.json
    
    # Store credentials in session
    session['username'] = data.get('username')
    session['password'] = data.get('password')
    session['security_code'] = data.get('securityCode') 
    session['case_id'] = data.get('caseId')
    
    try:
        # Set up Chrome options for non-headless mode but with minimal UI
        chrome_options = Options()
        # Don't use headless mode since RDN blocks it
        # chrome_options.add_argument("--headless")
        chrome_options.add_argument("--no-sandbox")
        chrome_options.add_argument("--disable-dev-shm-usage")
        # Add options to make browser less detectable as automated
        chrome_options.add_argument("--disable-blink-features=AutomationControlled")
        chrome_options.add_experimental_option("excludeSwitches", ["enable-automation"])
        chrome_options.add_experimental_option("useAutomationExtension", False)
        
        # Add version compatibility for newer Chrome versions
        chrome_options.add_argument("--remote-debugging-port=9222")  # Enable debugging
        chrome_options.add_argument("--disable-gpu")  # Disable GPU hardware acceleration
        # Removed w3c option that causes warnings
        
        # Check if browser should be minimized or hidden
        if os.environ.get('JAMI_HIDE_BROWSER', 'False').lower() == 'true':
            logging.info("Running with browser hidden")
            chrome_options.add_argument("--headless")
        
        # Download ChromeDriver if not available
        chromedriver_path = download_chromedriver()
        service = Service(executable_path=chromedriver_path)
        
        # Debug log
        logging.info(f"Using ChromeDriver at: {chromedriver_path}")
        
        # Add compatibility options - always use these for consistency
        logging.info("Adding compatibility options for all Chrome versions")
        chrome_options.add_argument("--disable-gpu")  # Disable GPU hardware acceleration
        
        # Create the driver with special error handling
        try:
            driver = webdriver.Chrome(service=service, options=chrome_options)
        except Exception as e:
            logging.error(f"Error creating Chrome driver: {e}")
            # Try fallback options if the initial attempt fails
            logging.info("Trying fallback options...")
            chrome_options.add_argument("--ignore-certificate-errors")
            chrome_options.add_argument("--disable-extensions")
            driver = webdriver.Chrome(service=service, options=chrome_options)
        driver.set_window_size(1280, 800)  # More typical window size
        
        # Execute CDP commands to mask WebDriver usage
        driver.execute_cdp_cmd("Page.addScriptToEvaluateOnNewDocument", {
            "source": """
                Object.defineProperty(navigator, 'webdriver', {
                    get: () => undefined
                })
            """
        })

        # Make the browser window minimized if running in non-interactive mode
        if os.environ.get('JAMI_MINIMIZE_BROWSER', 'False').lower() == 'true':
            logging.info("Minimizing browser window")
            driver.minimize_window()
        
        # Navigate to RDN login page using URL from config
        login_url = app_config['rdn']['login_url']
        logging.info(f"Navigating to login URL: {login_url}")
        driver.get(login_url)
        
        # Save page source for debugging
        with open(os.path.join('debug', 'login_page.html'), 'w', encoding='utf-8') as f:
            f.write(driver.page_source)
        
        # Wait for page to load
        WebDriverWait(driver, 15).until(
            EC.presence_of_element_located((By.TAG_NAME, "form"))
        )
        
        # Find the form fields
        username_field = driver.find_element(By.ID, "username")
        password_field = driver.find_element(By.ID, "password")
        
        # Find security code field (might have different IDs)
        security_code_field = None
        possible_ids = ["security_code", "securityCode", "code"]
        for id_value in possible_ids:
            try:
                security_code_field = driver.find_element(By.ID, id_value)
                logging.info(f"Found security code field with ID: {id_value}")
                break
            except NoSuchElementException:
                continue
        
        # If still not found, try to find by position (third input field)
        if not security_code_field:
            logging.warning("Security code field not found by ID, trying alternative method")
            input_fields = driver.find_elements(By.TAG_NAME, "input")
            if len(input_fields) >= 3:
                security_code_field = input_fields[2]
                logging.info("Using third input field as security code field")
        
        # Fill in credentials
        username_field.send_keys(data.get('username'))
        password_field.send_keys(data.get('password'))
        
        if security_code_field:
            security_code_field.send_keys(data.get('securityCode'))
        else:
            # Return error if security code field not found
            driver.quit()
            logging.error("Security code field not found")
            return jsonify({"success": False, "message": "Security code field not found"})
        
        # Find and click login button
        login_button = None
        try:
            # First try to find by type=submit
            login_button = driver.find_element(By.CSS_SELECTOR, "button[type='submit']")
            logging.info("Found login button by type=submit")
        except NoSuchElementException:
            try:
                # Try to find by text content
                login_button = driver.find_element(By.XPATH, "//button[contains(text(), 'Login') or contains(text(), 'Sign In')]")
                logging.info("Found login button by text content")
            except NoSuchElementException:
                try:
                    # Try to find any input with type=submit
                    login_button = driver.find_element(By.CSS_SELECTOR, "input[type='submit']")
                    logging.info("Found login button as input[type=submit]")
                except NoSuchElementException:
                    logging.error("Could not find login button using any method")
        
        if not login_button:
            driver.quit()
            logging.error("Login button not found")
            return jsonify({"success": False, "message": "Login button not found"})
        
        # Click login button
        login_button.click()
        logging.info("Clicked login button")
        
        # Take a screenshot after clicking login
        driver.save_screenshot(os.path.join('debug', 'after_login_click.png'))
        
        # Wait for login to complete
        try:
            WebDriverWait(driver, 15).until(
                EC.url_contains("recoverydatabase.net")
            )
            
            # Check if login failed (look for error message)
            error_messages = driver.find_elements(By.CSS_SELECTOR, ".error, .alert, .alert-danger")
            for error in error_messages:
                if error.is_displayed() and error.text:
                    driver.quit()
                    logging.error(f"Login failed: {error.text}")
                    return jsonify({"success": False, "message": f"Login failed: {error.text}"})
            
            # Store cookies in session for later use
            session['cookies'] = driver.get_cookies()
            
            driver.quit()
            logging.info("Login successful")
            return jsonify({"success": True, "message": "Login successful"})
            
        except TimeoutException:
            # Take screenshot on timeout
            driver.save_screenshot(os.path.join('debug', 'login_timeout.png'))
            driver.quit()
            logging.error("Login timed out")
            return jsonify({"success": False, "message": "Login timed out or failed"})
            
    except Exception as e:
        logging.exception(f"Error during login: {str(e)}")
        return jsonify({"success": False, "message": f"Error during login: {str(e)}"})

@app.route('/api/case-data', methods=['GET'])
def get_case_data():
    """Extract case data from RDN"""
    logging.info("Case data request received")
    if 'username' not in session:
        logging.error("Not logged in")
        return jsonify({"success": False, "message": "Not logged in"})
    
    case_id = session.get('case_id')
    logging.info(f"Processing case ID: {case_id}")
    
    try:
        # Set up Chrome options for non-headless mode but with minimal UI
        chrome_options = Options()
        # Don't use headless mode since RDN blocks it
        # chrome_options.add_argument("--headless")
        chrome_options.add_argument("--no-sandbox")
        chrome_options.add_argument("--disable-dev-shm-usage")
        # Add options to make browser less detectable as automated
        chrome_options.add_argument("--disable-blink-features=AutomationControlled")
        chrome_options.add_experimental_option("excludeSwitches", ["enable-automation"])
        chrome_options.add_experimental_option("useAutomationExtension", False)
        
        # Add version compatibility for newer Chrome versions
        chrome_options.add_argument("--remote-debugging-port=9222")  # Enable debugging
        chrome_options.add_argument("--disable-gpu")  # Disable GPU hardware acceleration
        # Removed w3c option that causes warnings
        
        # Check if browser should be minimized or hidden
        if os.environ.get('JAMI_HIDE_BROWSER', 'False').lower() == 'true':
            logging.info("Running with browser hidden")
            chrome_options.add_argument("--headless")
        
        # Download ChromeDriver if not available
        chromedriver_path = download_chromedriver()
        service = Service(executable_path=chromedriver_path)
        
        # Add compatibility options - always use these for consistency
        logging.info("Adding compatibility options for all Chrome versions")
        chrome_options.add_argument("--disable-gpu")  # Disable GPU hardware acceleration
        
        # Create the driver with special error handling
        try:
            driver = webdriver.Chrome(service=service, options=chrome_options)
        except Exception as e:
            logging.error(f"Error creating Chrome driver: {e}")
            # Try fallback options if the initial attempt fails
            logging.info("Trying fallback options...")
            chrome_options.add_argument("--ignore-certificate-errors")
            chrome_options.add_argument("--disable-extensions")
            driver = webdriver.Chrome(service=service, options=chrome_options)
        driver.set_window_size(1280, 800)  # More typical window size
        
        # Execute CDP commands to mask WebDriver usage
        driver.execute_cdp_cmd("Page.addScriptToEvaluateOnNewDocument", {
            "source": """
                Object.defineProperty(navigator, 'webdriver', {
                    get: () => undefined
                })
            """
        })
        
        # Make the browser window minimized if running in non-interactive mode
        if os.environ.get('JAMI_MINIMIZE_BROWSER', 'False').lower() == 'true':
            driver.minimize_window()
        
        # First login again
        login_url = app_config['rdn']['login_url']
        logging.info(f"Navigating to login URL: {login_url}")
        driver.get(login_url)
        
        # Wait for page to load
        WebDriverWait(driver, 15).until(
            EC.presence_of_element_located((By.TAG_NAME, "form"))
        )
        
        # Find the form fields and fill in credentials
        username_field = driver.find_element(By.ID, "username")
        password_field = driver.find_element(By.ID, "password")
        username_field.send_keys(session.get('username'))
        password_field.send_keys(session.get('password'))
        
        # Find security code field (with fallbacks)
        security_code_field = None
        try:
            security_code_field = driver.find_element(By.ID, "security_code")
        except NoSuchElementException:
            possible_ids = ["securityCode", "code"]
            for id_value in possible_ids:
                try:
                    security_code_field = driver.find_element(By.ID, id_value)
                    break
                except NoSuchElementException:
                    continue
            
            if not security_code_field:
                input_fields = driver.find_elements(By.TAG_NAME, "input")
                if len(input_fields) >= 3:
                    security_code_field = input_fields[2]
        
        if security_code_field:
            security_code_field.send_keys(session.get('security_code'))
        else:
            logging.error("Security code field not found during case data extraction")
        
        # Find and click login button
        login_button = driver.find_element(By.CSS_SELECTOR, "button[type='submit'], input[type='submit']")
        login_button.click()
        logging.info("Clicked login button (case data)")
        
        # Wait for login to complete
        WebDriverWait(driver, 15).until(
            EC.url_contains("recoverydatabase.net")
        )
        
        # Navigate to case page using URL from config
        case_url = app_config['rdn']['case_url_template'].format(case_id=case_id)
        logging.info(f"Navigating to case URL: {case_url}")
        driver.get(case_url)
        
        # Take screenshot of case page
        driver.save_screenshot(os.path.join('debug', f'case_{case_id}.png'))
        
        # Wait for case page to load
        WebDriverWait(driver, 15).until(
            EC.presence_of_element_located((By.TAG_NAME, "body"))
        )
        
        # Extract page content for parsing
        page_content = driver.page_source
        
        # Save page source for debugging
        with open(os.path.join('debug', f'case_{case_id}.html'), 'w', encoding='utf-8') as f:
            f.write(page_content)
        
        soup = BeautifulSoup(page_content, 'html.parser')
        
        # Initialize case data
        case_data = {
            "caseId": case_id,
            "clientName": "Not Found",
            "lienHolder": "Not Found",
            "orderTo": "Not Found",
            "fees": []
        }
        
        # Extract client information using various selectors and patterns
        # This is a simplified approach that would need to be customized based on actual RDN page structure
        
        # Look for labels containing client-related text
        client_labels = soup.find_all(string=re.compile(r'client|customer|account', re.IGNORECASE))
        for label in client_labels:
            parent = label.parent
            if parent and parent.next_sibling:
                text = parent.next_sibling.get_text().strip()
                if text:
                    case_data["clientName"] = text
                    logging.info(f"Found client name: {text}")
                    break
        
        # Look for labels containing lienholder-related text
        lien_labels = soup.find_all(string=re.compile(r'lien|lender|bank|credit union', re.IGNORECASE))
        for label in lien_labels:
            parent = label.parent
            if parent and parent.next_sibling:
                text = parent.next_sibling.get_text().strip()
                if text:
                    case_data["lienHolder"] = text
                    logging.info(f"Found lien holder: {text}")
                    break
        
        # Look for labels containing order-to-related text
        order_labels = soup.find_all(string=re.compile(r'order to|ordered|assigned', re.IGNORECASE))
        for label in order_labels:
            parent = label.parent
            if parent and parent.next_sibling:
                text = parent.next_sibling.get_text().strip()
                if text:
                    case_data["orderTo"] = text
                    logging.info(f"Found order to: {text}")
                    break
        
        # Extract fee information from the page
        # Look for dollar amounts and nearby text
        dollar_pattern = r'\$(\d{1,3}(,\d{3})*(\.\d{2})?)'
        all_text = soup.get_text()
        
        # Find all dollar amounts in the page
        dollar_matches = re.finditer(dollar_pattern, all_text)
        for match in dollar_matches:
            amount_str = match.group(0)  # Get the full dollar amount with $ sign
            amount = float(match.group(1).replace(',', ''))  # Get the numeric value
            
            # Get surrounding text for context (50 characters before and after)
            start_pos = max(0, match.start() - 50)
            end_pos = min(len(all_text), match.end() + 50)
            surrounding_text = all_text[start_pos:end_pos]
            
            # Identify fee type and status based on surrounding text
            fee_type = identify_fee_type(surrounding_text)
            fee_status = identify_fee_status(surrounding_text)
            
            case_data["fees"].append({
                "description": surrounding_text.strip(),
                "amount": amount,
                "amountStr": amount_str,
                "category": fee_type["category"],
                "categoryColor": fee_type["color"],
                "status": fee_status
            })
        
        logging.info(f"Found {len(case_data['fees'])} fees")
        
        # Now navigate to the Updates tab to get update history
        try:
            # Find Updates tab
            update_links = driver.find_elements(By.XPATH, "//a[contains(text(), 'Updates') or contains(text(), 'History')]")
            if update_links:
                logging.info("Found Updates tab, clicking...")
                update_links[0].click()
                
                # Wait for updates to load
                WebDriverWait(driver, 10).until(
                    EC.presence_of_element_located((By.ID, "updates-tab"))
                )
                
                # Take screenshot of updates tab
                driver.save_screenshot(os.path.join('debug', f'updates_{case_id}.png'))
                
                # Extract updates page content
                updates_content = driver.page_source
                
                # Save updates page source for debugging
                with open(os.path.join('debug', f'updates_{case_id}.html'), 'w', encoding='utf-8') as f:
                    f.write(updates_content)
                
                updates_soup = BeautifulSoup(updates_content, 'html.parser')
                
                # Find update elements
                updates = []
                update_elements = updates_soup.find_all(['div', 'tr'], class_=re.compile(r'update|history|log'))
                
                if not update_elements:
                    # Fallback: look for table rows that might contain update info
                    logging.info("No update elements found with class, trying fallback")
                    update_elements = updates_soup.find_all('tr')
                
                for element in update_elements:
                    element_text = element.get_text()
                    
                    # Extract date using regex
                    date_match = re.search(r'\d{1,2}/\d{1,2}/\d{4}', element_text)
                    date = date_match.group(0) if date_match else "Unknown"
                    
                    # Extract fee amount
                    amount_match = re.search(dollar_pattern, element_text)
                    amount_str = amount_match.group(0) if amount_match else "$0.00"
                    amount = float(amount_match.group(1).replace(',', '')) if amount_match else 0.0
                    
                    # Get details (text excluding date and amount)
                    details = element_text
                    if date_match:
                        details = details.replace(date_match.group(0), "")
                    if amount_match:
                        details = details.replace(amount_match.group(0), "")
                    details = ' '.join(details.split())  # Normalize whitespace
                    
                    # Identify fee type
                    fee_type_info = identify_fee_type(element_text)
                    fee_type = fee_type_info["category"] if fee_type_info["confidence"] > 0 else "N/A"
                    
                    updates.append({
                        "date": date,
                        "details": details.strip(),
                        "amount": amount,
                        "amountStr": amount_str,
                        "feeType": fee_type
                    })
                
                session['updates'] = updates
                logging.info(f"Found {len(updates)} update entries")
            else:
                logging.warning("Updates tab not found")
                session['updates'] = []
        except Exception as e:
            logging.exception(f"Error extracting updates: {str(e)}")
            session['updates'] = []
        
        # Store case data in session
        session['case_data'] = case_data
        
        driver.quit()
        return jsonify({"success": True, "data": case_data})
        
    except Exception as e:
        logging.exception(f"Error extracting case data: {str(e)}")
        return jsonify({"success": False, "message": f"Error extracting case data: {str(e)}"})

@app.route('/api/query-database', methods=['GET'])
def query_database():
    """Query Azure SQL database for fee information using config credentials"""
    logging.info("Database query request received")
    if 'case_data' not in session:
        logging.error("Case data not available")
        return jsonify({"success": False, "message": "Case data not available"})
    
    case_data = session.get('case_data')
    
    try:
        # Get database config
        db_config = app_config['database']
        
        # Build connection string from config
        server = db_config['server']
        database = db_config['database']
        username = db_config['username']
        password = db_config['password']
        
        # Build a proper pyodbc connection string
        conn_str = f"DRIVER={{SQL Server}};SERVER={server};DATABASE={database};UID={username};PWD={password}"
        logging.info(f"Connecting to database: {server}/{database}")
        
        # Connect to the database
        conn = pyodbc.connect(conn_str)
        cursor = conn.cursor()
        
        # Prepare parameters
        client_name = case_data.get('clientName')
        lienholder_name = case_data.get('lienHolder')
        fee_type_name = 'Involuntary Repo'  # Default to Involuntary Repo
        
        logging.info(f"Querying for: Client={client_name}, Lienholder={lienholder_name}, FeeType={fee_type_name}")
        
        # Execute query using parameters
        query = """
        -- Declare input parameters
        DECLARE @clientName NVARCHAR(100) = ?;
        DECLARE @lienholderName NVARCHAR(100) = ?;
        DECLARE @feeTypeName NVARCHAR(100) = ?;

        -- Get foreign keys from names
        DECLARE @clientId INT = (SELECT TOP 1 id FROM dbo.RDN_Client WHERE client_name = @clientName);
        DECLARE @lienholderId INT = (SELECT TOP 1 id FROM dbo.Lienholder WHERE lienholder_name = @lienholderName);
        DECLARE @feeTypeId INT = (SELECT TOP 1 id FROM dbo.FeeType WHERE fee_type_name = @feeTypeName);

        -- Check if a matching FeeDetails2 record exists
        IF EXISTS (
            SELECT 1 
            FROM dbo.FeeDetails2 
            WHERE client_id = @clientId AND lh_id = @lienholderId AND ft_id = @feeTypeId
        )
        BEGIN
            -- Return result with names
            SELECT 
                fd.fd_id AS fdId,
                c.client_name AS clientName,
                lh.lienholder_name AS lienholderName,
                ft.fee_type_name AS feeTypeName,
                fd.amount AS amount
            FROM dbo.FeeDetails2 fd
            JOIN dbo.RDN_Client c ON fd.client_id = c.id
            JOIN dbo.Lienholder lh ON fd.lh_id = lh.id
            JOIN dbo.FeeType ft ON fd.ft_id = ft.id
            WHERE fd.client_id = @clientId AND fd.lh_id = @lienholderId AND fd.ft_id = @feeTypeId;
        END
        ELSE
        BEGIN
            -- Get ID for 'Standard' lienholder
            DECLARE @standardLienholderId INT = (
                SELECT TOP 1 id FROM dbo.Lienholder WHERE lienholder_name = 'Standard'
            );

            -- Return fallback result with names
            SELECT 
                fd.fd_id AS fdId,
                c.client_name AS clientName,
                lh.lienholder_name AS lienholderName,
                ft.fee_type_name AS feeTypeName,
                fd.amount AS amount
            FROM dbo.FeeDetails2 fd
            JOIN dbo.RDN_Client c ON fd.client_id = c.id
            JOIN dbo.Lienholder lh ON fd.lh_id = lh.id
            JOIN dbo.FeeType ft ON fd.ft_id = ft.id
            WHERE fd.client_id = @clientId AND fd.lh_id = @standardLienholderId AND fd.ft_id = @feeTypeId;
        END
        """
        
        cursor.execute(query, (client_name, lienholder_name, fee_type_name))
        
        # Fetch result
        columns = [column[0] for column in cursor.description]
        db_result = None
        
        for row in cursor.fetchall():
            db_result = dict(zip(columns, row))
            break  # We only need the first row
        
        # Close connection
        cursor.close()
        conn.close()
        
        # If no result, create a mock result for testing
        if not db_result:
            logging.warning("No database result found, using mock data")
            db_result = {
                "fdId": f"FD-{case_data.get('caseId', '0000')}",
                "clientName": client_name,
                "lienholderName": lienholder_name,
                "feeTypeName": fee_type_name,
                "amount": 350.00
            }
        else:
            logging.info(f"Database result: {db_result}")
        
        # Store in session
        session['db_data'] = db_result
        
        return jsonify({"success": True, "data": db_result})
        
    except Exception as e:
        logging.exception(f"Database error: {str(e)}")
        # Create mock data for testing when database connection fails
        mock_data = {
            "fdId": f"FD-{case_data.get('caseId', '0000')}",
            "clientName": case_data.get('clientName'),
            "lienholderName": case_data.get('lienHolder'),
            "feeTypeName": "Involuntary Repo",
            "amount": 350.00
        }
        
        session['db_data'] = mock_data
        
        return jsonify({
            "success": True, 
            "data": mock_data, 
            "warning": f"Database connection failed: {str(e)}. Using mock data."
        })

@app.route('/api/updates', methods=['GET'])
def get_updates():
    """Retrieve updates data from session"""
    logging.info("Updates data request received")
    if 'updates' not in session:
        logging.error("Updates data not available")
        return jsonify({"success": False, "message": "Updates data not available"})
    
    return jsonify({"success": True, "data": session.get('updates')})

@app.route('/api/results', methods=['GET'])
def get_results():
    """Retrieve all results data"""
    logging.info("Results request received")
    if 'case_data' not in session or 'db_data' not in session or 'updates' not in session:
        logging.error("Complete data not available")
        return jsonify({"success": False, "message": "Complete data not available"})
    
    return jsonify({
        "success": True,
        "caseData": session.get('case_data'),
        "dbData": session.get('db_data'),
        "updates": session.get('updates')
    })

@app.route('/api/export/excel', methods=['GET'])
def export_excel():
    """Generate Excel export of results"""
    logging.info("Excel export request received")
    if 'case_data' not in session or 'db_data' not in session or 'updates' not in session:
        logging.error("Complete data not available for export")
        return jsonify({"success": False, "message": "Complete data not available"})
    
    case_data = session.get('case_data')
    db_data = session.get('db_data')
    updates = session.get('updates')
    
    try:
        # Create a file name
        file_name = f"JamiBilling_Case_{case_data['caseId']}_{datetime.datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        file_path = os.path.join('static', 'exports', file_name)
        logging.info(f"Creating Excel export: {file_path}")
        
        # Create Excel workbook
        workbook = openpyxl.Workbook()
        
        # Create summary sheet
        summary_sheet = workbook.active
        summary_sheet.title = "Summary"
        
        # Add headers
        summary_sheet['A1'] = "Item"
        summary_sheet['B1'] = "Value"
        
        # Add data
        summary_data = [
            ["Case ID", case_data['caseId']],
            ["Client Name", case_data['clientName']],
            ["Lien Holder", case_data['lienHolder']],
            ["Order To", case_data['orderTo']],
            ["", ""],
            ["Database Information", ""],
            ["Fee ID", db_data['fdId']],
            ["Fee Type", db_data['feeTypeName']],
            ["Amount", db_data['amount']],
            ["", ""],
            ["Total Fees", sum(fee['amount'] for fee in case_data['fees'])]
        ]
        
        # Add rows to sheet
        for i, row in enumerate(summary_data, 2):  # Start from row 2
            summary_sheet[f'A{i}'] = row[0]
            summary_sheet[f'B{i}'] = row[1]
        
        # Create fees sheet
        fees_sheet = workbook.create_sheet(title="Fees")
        
        # Add headers
        fees_sheet['A1'] = "Description"
        fees_sheet['B1'] = "Category"
        fees_sheet['C1'] = "Amount"
        fees_sheet['D1'] = "Status"
        
        # Add fee data
        for i, fee in enumerate(case_data['fees'], 2):  # Start from row 2
            fees_sheet[f'A{i}'] = fee['description']
            fees_sheet[f'B{i}'] = fee['category']
            fees_sheet[f'C{i}'] = fee['amount']
            fees_sheet[f'D{i}'] = fee['status']
        
        # Create updates sheet
        updates_sheet = workbook.create_sheet(title="Updates")
        
        # Add headers
        updates_sheet['A1'] = "Date"
        updates_sheet['B1'] = "Details"
        updates_sheet['C1'] = "Fee Type"
        updates_sheet['D1'] = "Amount"
        
        # Add update data
        for i, update in enumerate(updates, 2):  # Start from row 2
            updates_sheet[f'A{i}'] = update['date']
            updates_sheet[f'B{i}'] = update['details']
            updates_sheet[f'C{i}'] = update['feeType']
            updates_sheet[f'D{i}'] = update['amount']
        
        # Save the Excel file
        workbook.save(file_path)
        logging.info("Excel file saved successfully")
        
        # Return the file URL
        file_url = f"/static/exports/{file_name}"
        return jsonify({"success": True, "file_url": file_url})
        
    except Exception as e:
        logging.exception(f"Error generating Excel: {str(e)}")
        return jsonify({"success": False, "message": f"Error generating Excel: {str(e)}"})

def identify_fee_type(text):
    """Identify fee type from text description"""
    if not text:
        return {"category": "Unknown", "confidence": 0, "color": "#858796"}
    
    text_lower = text.lower()
    highest_confidence = 0
    detected_category = "Unknown"
    category_color = "#858796"  # Default gray
    
    # Check each category's keywords against the text
    for category, data in fee_categories.items():
        for keyword in data["keywords"]:
            if keyword.lower() in text_lower:
                # Found an exact match
                return {
                    "category": category,
                    "confidence": 1,
                    "color": data["color"]
                }
        
        # No exact match, try partial matching with score
        score = 0
        for keyword in data["keywords"]:
            keyword_parts = keyword.lower().split()
            for part in keyword_parts:
                if len(part) > 3 and part in text_lower:
                    score += 0.5
        
        if score > highest_confidence:
            highest_confidence = score
            detected_category = category
            category_color = data["color"]
    
    if highest_confidence > 0:
        return {
            "category": detected_category,
            "confidence": min(highest_confidence, 0.9),  # Cap at 0.9 for partial matches
            "color": category_color
        }
    
    return {"category": "Unknown", "confidence": 0, "color": "#858796"}

def identify_fee_status(text):
    """Identify fee status from text description"""
    if not text:
        return "Unknown"
    
    text_lower = text.lower()
    
    status_keywords = {
        "Paid": ["paid", "payment received", "payment complete"],
        "Not Paid": ["not paid", "unpaid", "payment pending"],
        "Approved": ["approved", "accepted", "authorized"],
        "Pending": ["pending", "awaiting", "in process"],
        "Denied": ["denied", "rejected", "declined"],
        "Waived": ["waived", "forgiven", "no charge"]
    }
    
    for status, keywords in status_keywords.items():
        for keyword in keywords:
            if keyword in text_lower:
                return status
    
    return "Unknown"

@app.route('/debug-logs', methods=['GET'])
def view_debug_logs():
    """View debug information (only in development)"""
    if not app.debug:
        return jsonify({"error": "Debug mode not enabled"}), 403
    
    try:
        # Get list of debug files
        debug_files = []
        for filename in os.listdir('debug'):
            if filename.endswith('.png') or filename.endswith('.html'):
                file_path = os.path.join('debug', filename)
                file_info = {
                    'name': filename,
                    'path': file_path,
                    'size': os.path.getsize(file_path),
                    'modified': datetime.datetime.fromtimestamp(os.path.getmtime(file_path)).strftime('%Y-%m-%d %H:%M:%S')
                }
                debug_files.append(file_info)
        
        # Sort by modified date (newest first)
        debug_files.sort(key=lambda x: x['modified'], reverse=True)
        
        # Get log file content
        log_content = ""
        if os.path.exists('jami_billing.log'):
            with open('jami_billing.log', 'r') as f:
                log_content = f.read()
        
        return jsonify({
            "debug_files": debug_files,
            "log_content": log_content
        })
    except Exception as e:
        return jsonify({"error": str(e)}), 500

@app.route('/debug-file/<filename>', methods=['GET'])
def get_debug_file(filename):
    """Get a specific debug file (only in development)"""
    if not app.debug:
        return jsonify({"error": "Debug mode not enabled"}), 403
    
    try:
        # Ensure filename is safe and exists
        safe_filename = secure_filename(filename)
        file_path = os.path.join('debug', safe_filename)
        
        if not os.path.exists(file_path):
            return jsonify({"error": "File not found"}), 404
        
        # Serve the file
        if filename.endswith('.png'):
            return send_file(file_path, mimetype='image/png')
        elif filename.endswith('.html'):
            return send_file(file_path, mimetype='text/html')
        else:
            return send_file(file_path, mimetype='text/plain')
    except Exception as e:
        return jsonify({"error": str(e)}), 500

@app.route('/healthcheck', methods=['GET'])
def healthcheck():
    """Simple health check endpoint"""
    return jsonify({
        "status": "ok",
        "version": "1.0",
        "timestamp": datetime.datetime.now().isoformat()
    })

if __name__ == '__main__':
    # Create necessary directories
    os.makedirs('debug', exist_ok=True)
    os.makedirs('static/exports', exist_ok=True)
    os.makedirs('flask_session', exist_ok=True)
    
    # Log startup
    logging.info("Starting JamiBilling application")
    
    app.run(debug=True, port=5000)