"""
JamiBilling - RDN Fee Scraper Application
Playwright version with improved browser automation
"""

import os
import json
import re
import csv
import datetime
import io
import logging
import time
import asyncio
import decimal
import traceback
import pypyodbc as pyodbc
from flask import Flask, render_template, request, jsonify, session, send_file
from flask_session import Session
from playwright.async_api import async_playwright
from bs4 import BeautifulSoup
from werkzeug.utils import secure_filename
import openpyxl

# Configure logging
logging.basicConfig(
    filename='jami_billing.log',
    level=logging.INFO,
    format='%(asctime)s - %(name)s - %(levelname)s - %(message)s'
)

# Initialize Flask app
app = Flask(__name__)
app.config["SECRET_KEY"] = os.urandom(24)
app.config["SESSION_TYPE"] = "filesystem"
app.config["SESSION_PERMANENT"] = False
app.config["SESSION_FILE_DIR"] = "flask_session"
Session(app)

# Load application config
try:
    with open('config.json', 'r') as f:
        app_config = json.load(f)
    logging.info("Config loaded successfully")
except Exception as e:
    logging.error(f"Error loading config: {str(e)}")
    app_config = {
        "database": {
            "provider": "SQLOLEDB",
            "server": "localhost",
            "database": "RDN_Billing",
            "username": "admin",
            "password": "password"
        },
        "rdn": {
            "login_url": "https://secureauth.recoverydatabase.net/public/login",
            "case_url_template": "https://app.recoverydatabase.net/alpha_rdn/module/default/case2/?case_id={case_id}"
        }
    }

# Load fee categories from the JSON file
try:
    with open('backend/fee_categories.json', 'r') as f:
        fee_categories = json.load(f)
    logging.info("Fee categories loaded successfully")
except Exception as e:
    logging.error(f"Error loading fee categories: {str(e)}")
    fee_categories = {
        "Storage": {
            "keywords": ["storage fee", "daily storage", "impound fee", "lot fee"],
            "color": "#4e73df"
        },
        "Repossession": {
            "keywords": ["repo fee", "recovery fee", "tow fee", "involuntary repo"],
            "color": "#e74a3b"
        },
        "Other": {
            "keywords": ["other"],
            "color": "#858796"
        }
    }

# Ensure required directories exist
os.makedirs('flask_session', exist_ok=True)
os.makedirs(os.path.join('static', 'exports'), exist_ok=True)
os.makedirs('debug', exist_ok=True)

# Global event loop for asyncio
loop = None

def get_event_loop():
    """Get or create the event loop for async operations"""
    global loop
    if loop is None or loop.is_closed():
        loop = asyncio.new_event_loop()
        asyncio.set_event_loop(loop)
    return loop

@app.route('/')
def index():
    """Render the main application page"""
    logging.info("Rendering main page")
    return render_template('index.html')

@app.route('/api/login', methods=['POST'])
def login():
    """Handle login to RDN using Playwright"""
    logging.info("Login request received")
    data = request.json
    
    # Store credentials in session
    session['username'] = data.get('username')
    session['password'] = data.get('password')
    session['security_code'] = data.get('securityCode') 
    session['case_id'] = data.get('caseId')
    
    # Add support for multi-factor authentication
    session['is_second_step'] = data.get('is_second_step', False)
    session['verification_code'] = data.get('verificationCode', '')
    
    try:
        # Run the async login function in the event loop
        loop = get_event_loop()
        success, message = loop.run_until_complete(async_login(data))
        
        # If this looks like it might be waiting for a second factor, inform client
        if success is False and "multi-factor" in message.lower():
            return jsonify({
                "success": False, 
                "requires_2fa": True,
                "message": message
            })
        
        return jsonify({"success": success, "message": message})
            
    except Exception as e:
        logging.exception(f"Error during login: {str(e)}")
        return jsonify({"success": False, "message": f"Error during login: {str(e)}"})

async def async_login(data):
    """Async function to handle Playwright browser automation for login"""
    async with async_playwright() as p:
        # Launch browser - Playwright's chromium has better automation detection avoidance
        logging.info("Launching Playwright browser")
        browser = await p.chromium.launch(
            headless=os.environ.get('JAMI_HIDE_BROWSER', 'False').lower() == 'true'
        )
        
        try:
            # Create a new browser context with specific options
            context = await browser.new_context(
                viewport={"width": 1280, "height": 800},
                user_agent="Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/136.0.0.0 Safari/537.36"
            )
            
            # Create a new page
            page = await context.new_page()
            
            # Navigate to RDN login page using URL from config
            login_url = app_config['rdn']['login_url']
            logging.info(f"Navigating to login URL: {login_url}")
            await page.goto(login_url)
            
            # Save page source for debugging
            login_page_content = await page.content()
            with open(os.path.join('debug', 'login_page.html'), 'w', encoding='utf-8') as f:
                f.write(login_page_content)
            
            # Take screenshot
            await page.screenshot(path=os.path.join('debug', 'login_page.png'))
            
            # Save screenshot before filling
            await page.screenshot(path=os.path.join('debug', 'before_fill.png'))
            
            # Fill form fields using direct selectors for maximum performance
            logging.info("Filling login form with optimized approach")
            
            # Use common selectors for login forms
            selectors = [
                # Username selectors
                {"field": "username", "selectors": [
                    "#username", "[name=username]", "input[type=text]:nth-of-type(1)",
                    "form input[type=text]:first-of-type", "input:not([type=hidden]):nth-of-type(1)"
                ]},
                # Password selectors
                {"field": "password", "selectors": [
                    "#password", "[name=password]", "input[type=password]",
                    "form input[type=password]", "input:not([type=hidden]):nth-of-type(2)"
                ]},
                # Security code selectors
                {"field": "securityCode", "selectors": [
                    "#security_code", "#securityCode", "#code",
                    "[name=security_code]", "[name=securityCode]", "[name=code]",
                    "input:not([type=hidden]):nth-of-type(3)"
                ]}
            ]
            
            # Fill each field using the first working selector
            security_code_filled = False
            for field_info in selectors:
                field_name = field_info["field"]
                field_value = data.get(field_name)
                
                if not field_value:
                    continue
                    
                # Try each selector until one works
                for selector in field_info["selectors"]:
                    try:
                        # Use a short timeout to avoid waiting too long for each selector
                        element = await page.wait_for_selector(selector, timeout=50)
                        if element:
                            await element.fill(field_value)
                            logging.info(f"{field_name} filled using {selector}")
                            if field_name == "securityCode":
                                security_code_filled = True
                            break
                    except Exception:
                        continue
            
            # Fallback for security code if not filled
            if not security_code_filled and data.get('securityCode'):
                logging.warning("Security code not filled with selectors, trying fallback")
                try:
                    # Get all visible inputs
                    inputs = await page.query_selector_all("input:not([type=hidden])")
                    # If we have at least 3 inputs, use the third one for security code
                    if len(inputs) >= 3:
                        await inputs[2].fill(data.get('securityCode'))
                        logging.info("Security code filled using fallback method")
                        security_code_filled = True
                except Exception as e:
                    logging.error(f"Security code fallback failed: {str(e)}")
            
            # Take screenshot after filling the form
            await page.screenshot(path=os.path.join('debug', 'after_fill.png'))
            
            # Find and click login button - optimized approach with CAPTCHA handling
            logging.info("Checking for CAPTCHA before clicking login button")
            
            # Check for CAPTCHA and handle it
            try:
                captcha_frame = await page.query_selector("iframe[src*='recaptcha'], iframe[title*='reCAPTCHA'], div.g-recaptcha")
                if captcha_frame:
                    logging.info("CAPTCHA detected! Notifying user")
                    html_content = await page.content()
                    with open(os.path.join('debug', 'captcha_page.html'), 'w', encoding='utf-8') as f:
                        f.write(html_content)
                    await page.screenshot(path=os.path.join('debug', 'captcha_detected.png'))
                    
                    # If headless mode is on, we can't solve CAPTCHA
                    if os.environ.get('JAMI_HIDE_BROWSER', 'False').lower() == 'true':
                        await browser.close()
                        return False, "CAPTCHA detected - manual login required. Please set JAMI_HIDE_BROWSER=False to show the browser and solve the CAPTCHA."
                    else:
                        # Wait for user to solve CAPTCHA manually
                        logging.info("Browser is visible - waiting 60 seconds for user to solve CAPTCHA manually...")
                        await page.wait_for_timeout(60000)  # Wait 60 seconds for user to solve
                        logging.info("Continuing after CAPTCHA timeout")
            except Exception as e:
                logging.warning(f"Error checking for CAPTCHA: {str(e)}")
            
            # Try to submit form directly first
            try:
                logging.info("Attempting to submit form using JavaScript")
                await page.evaluate("document.querySelector('form').submit()")
                logging.info("Form submitted via JavaScript")
                button_clicked = True
            except Exception as e:
                logging.warning(f"JavaScript form submission failed: {str(e)}")
                
                # Fall back to clicking buttons
                logging.info("Trying button click approach")
                # Try multiple button selectors in order of likelihood
                button_selectors = [
                    "button.btn-success",  # The actual Login button class
                    "button:has-text('Login')", 
                    "button[type='submit']", 
                    "input[type='submit']",
                    "button:has-text('Sign In')",
                    "button:has-text('Submit')",
                    "input[type='button'][value='Login']",
                    "form button"
                ]
                
                button_clicked = False
                for selector in button_selectors:
                    try:
                        button = await page.wait_for_selector(selector, timeout=50)
                        if button:
                            await button.click()
                            logging.info(f"Clicked button using selector: {selector}")
                            button_clicked = True
                            break
                    except Exception:
                        continue
            
            # If no button found, press Enter in the last field
            if not button_clicked:
                try:
                    inputs = await page.query_selector_all("input:not([type=hidden])")
                    if inputs:
                        await inputs[-1].press("Enter")
                        logging.info("Pressed Enter on last input field")
                    else:
                        logging.error("No input fields found to press Enter")
                        await page.screenshot(path=os.path.join('debug', 'login_button_error.png'))
                        await browser.close()
                        return False, "No login button found and no input fields to press Enter"
                except Exception as e:
                    logging.error(f"Error submitting form: {str(e)}")
                    await page.screenshot(path=os.path.join('debug', 'login_button_error.png'))
                    await browser.close()
                    return False, f"Error submitting form: {str(e)}"
            
            # Wait for navigation and take screenshot
            try:
                # Wait for navigation using multiple possible patterns with longer timeout
                try:
                    # Try waiting for any successful navigation away from login
                    logging.info("Waiting for navigation after login...")
                    
                    # Just wait for load state instead of specific URL
                    await page.wait_for_load_state("domcontentloaded", timeout=10000)
                    
                    # Check if we're still on the login page
                    current_url = page.url
                    if "login" in current_url.lower():
                        logging.warning(f"Still on login page after button click: {current_url}")
                    else:
                        logging.info(f"Navigation successful: {current_url}")
                except Exception as e:
                    logging.warning(f"Navigation wait timed out, but proceeding: {str(e)}")
                await page.screenshot(path=os.path.join('debug', 'after_login.png'))
                
                # Check for error messages and analyze page
                current_url = page.url
                page_title = await page.title()
                
                logging.info(f"Current URL after login attempt: {current_url}")
                logging.info(f"Page title after login attempt: {page_title}")
                
                # Look for any error elements
                error_elements = await page.query_selector_all(".error, .alert, .alert-danger, .message-error, #error-message")
                for error_element in error_elements:
                    error_text = await error_element.text_content()
                    if error_text and error_text.strip():
                        logging.error(f"Login error found: {error_text.strip()}")
                        await browser.close()
                        return False, f"Login failed: {error_text.strip()}"
                
                # Consider the login successful if we're not on a page with "login" in the URL
                # and there are no error messages
                if "login" in current_url.lower():
                    # Check if there's a specific element indicating we're in the app
                    # This could be a dashboard element, menu, etc.
                    success_element = await page.query_selector("nav, .dashboard, .header-menu, .welcome-message")
                    if not success_element:
                        # Try one more click if we appear to be still on login page
                        try:
                            # Look for a multi-step login button
                            next_button = await page.query_selector("button:has-text('Continue'), button:has-text('Next'), button:has-text('Proceed')")
                            if next_button:
                                logging.info("Found a next/continue button, attempting to click it")
                                await next_button.click()
                                await page.wait_for_load_state("domcontentloaded", timeout=5000)
                                # Re-check URL after second click
                                new_url = page.url
                                if "login" not in new_url.lower():
                                    logging.info(f"Multi-step login successful, now at: {new_url}")
                                else:
                                    logging.warning("Still on login page after clicking continue button")
                                    # Save the current page for analysis
                                    html_content = await page.content()
                                    with open(os.path.join('debug', 'login_analysis.html'), 'w', encoding='utf-8') as f:
                                        f.write(html_content)
                                    await browser.close()
                                    return False, "Login appears to have failed - still on login page"
                            else:
                                # Save the current page for analysis
                                html_content = await page.content()
                                with open(os.path.join('debug', 'login_analysis.html'), 'w', encoding='utf-8') as f:
                                    f.write(html_content)
                                # Check if this appears to be a verification code / MFA page
                                verification_elements = await page.query_selector_all("input[name='verificationCode'], input[name='code'], input[placeholder*='code'], input[placeholder*='verification'], input[type='number']")
                                if verification_elements:
                                    # This looks like a multi-factor authentication page
                                    logging.info("Detected possible multi-factor authentication page")
                                    
                                    # If we have a verification code and this is marked as second step, try to use it
                                    if data.get('is_second_step') and data.get('verificationCode'):
                                        try:
                                            verification_code = data.get('verificationCode')
                                            logging.info(f"Attempting to submit verification code: {verification_code}")
                                            
                                            # Fill the verification code into the first field that looks right
                                            for element in verification_elements:
                                                await element.fill(verification_code)
                                                logging.info("Filled verification code")
                                                break
                                                
                                            # Look for a submit button
                                            verify_button = await page.query_selector("button:has-text('Verify'), button:has-text('Submit'), button:has-text('Continue'), button[type='submit']")
                                            if verify_button:
                                                await verify_button.click()
                                                logging.info("Clicked verification submit button")
                                                
                                                # Wait for result
                                                await page.wait_for_load_state("domcontentloaded", timeout=10000)
                                                
                                                # Check if we're past login now
                                                after_verify_url = page.url
                                                if "login" not in after_verify_url.lower():
                                                    logging.info(f"Two-factor authentication successful, now at: {after_verify_url}")
                                                    # Store cookies and continue
                                                    cookies = await context.cookies()
                                                    session['cookies'] = cookies
                                                    await page.screenshot(path=os.path.join('debug', 'after_2fa.png'))
                                                    await browser.close()
                                                    return True, "Login with two-factor authentication successful"
                                                else:
                                                    logging.error("Still on login page after verification attempt")
                                                    await page.screenshot(path=os.path.join('debug', 'failed_2fa.png'))
                                                    await browser.close() 
                                                    return False, "Verification code appears to be invalid"
                                        except Exception as e:
                                            logging.error(f"Error during verification code submission: {str(e)}")
                                            await browser.close()
                                            return False, f"Error during verification code submission: {str(e)}"
                                    else:
                                        # We need a verification code but don't have one
                                        logging.warning("Multi-factor authentication required")
                                        await page.screenshot(path=os.path.join('debug', 'needs_2fa.png'))
                                        await browser.close()
                                        return False, "Multi-factor authentication required - please provide a verification code"
                                else:
                                    # No verification element but we're still on login page
                                    logging.warning("No success elements found and still on login page")
                                    await browser.close()
                                    return False, "Login appears to have failed - possibly incorrect credentials"
                        except Exception as e:
                            logging.error(f"Error during multi-step login attempt: {str(e)}")
                            await browser.close()
                            return False, f"Login failed during multi-step process: {str(e)}"
                
                # If login successful, go directly to case page if case_id is available
                if data.get('caseId'):
                    try:
                        case_url = app_config['rdn']['case_url_template'].format(case_id=data.get('caseId'))
                        logging.info(f"Login successful, navigating directly to case URL: {case_url}")
                        
                        # Navigate to the case page
                        await page.goto(case_url)
                        
                        # Wait for case page to load
                        try:
                            await page.wait_for_load_state("networkidle", timeout=15000)  # Increased timeout
                            logging.info("Case page loaded successfully")
                        except Exception as e:
                            logging.warning(f"Timeout waiting for networkidle during navigation, continuing anyway: {str(e)}")
                            # Wait a reasonable time anyway
                            await page.wait_for_timeout(3000)
                        await page.screenshot(path=os.path.join('debug', f'direct_case_{data.get("caseId")}.png'))
                        
                        # Store cookies in session for later use
                        cookies = await context.cookies()
                        session['cookies'] = cookies
                        
                        logging.info("Login successful and case page loaded")
                        
                        # Extract basic page information to confirm successful load
                        title = await page.title()
                        logging.info(f"Case page title: {title}")
                        
                        await browser.close()
                        return True, "Login successful and case page loaded"
                    except Exception as e:
                        logging.error(f"Error navigating to case page: {str(e)}")
                        # Store cookies anyway since login was successful
                        cookies = await context.cookies()
                        session['cookies'] = cookies
                        await browser.close()
                        return True, "Login successful but case page navigation failed"
                
                # Store cookies in session for later use
                cookies = await context.cookies()
                session['cookies'] = cookies
                
                logging.info("Login successful")
                await browser.close()
                return True, "Login successful"
                
            except Exception as e:
                logging.error(f"Login timed out or failed: {str(e)}")
                await page.screenshot(path=os.path.join('debug', 'login_timeout.png'))
                await browser.close()
                return False, f"Login timed out or failed: {str(e)}"
                
        except Exception as e:
            logging.exception(f"Error during login: {str(e)}")
            await browser.close()
            return False, f"Error during login: {str(e)}"

@app.route('/api/case-data', methods=['GET'])
def get_case_data():
    """Extract case data from RDN"""
    logging.info("Case data request received")
    if 'username' not in session:
        logging.error("Not logged in")
        return jsonify({"success": False, "message": "Not logged in"})
    
    case_id = session.get('case_id')
    logging.info(f"Processing case ID: {case_id}")
    
    try:
        # Run the async case data extraction function in the event loop
        loop = get_event_loop()
        success, result = loop.run_until_complete(async_extract_case_data(case_id))
        
        if success:
            # Filter out fees with zero amounts before storing in session
            if "fees" in result["case_data"]:
                result["case_data"]["fees"] = [fee for fee in result["case_data"]["fees"] if fee.get('amount', 0) > 0]
                logging.info(f"Filtered fees to exclude zero amounts. Remaining fees: {len(result['case_data']['fees'])}")
                
            # Store case data and updates in session
            session['case_data'] = result["case_data"]
            session['updates'] = result["updates"]
            
            # Auto-fetch database data for this case
            logging.info("Case data extraction successful, auto-fetching database data")
            try:
                db_data = auto_fetch_database_fees(result["case_data"])
                if db_data:
                    logging.info(f"Auto-fetched database fee: ${db_data['amount']:.2f}")
                    # No need to store in session, auto_fetch_database_fees already does that
            except Exception as e:
                logging.error(f"Error during auto database fetch: {str(e)}")
                # This is non-blocking, so we continue even if db fetch fails
            
            return jsonify({"success": True, "data": result["case_data"]})
        else:
            return jsonify({"success": False, "message": result})
        
    except Exception as e:
        error_msg = str(e)
        logging.exception(f"Error extracting case data: {error_msg}")
        
        # Provide a more user-friendly message for timeout errors
        if "Timeout" in error_msg:
            return jsonify({
                "success": False, 
                "message": "The RDN system is taking too long to respond. Please try again or check RDN status.",
                "technical_error": error_msg
            })
        else:
            return jsonify({"success": False, "message": f"Error extracting case data: {error_msg}"})

async def async_extract_case_data(case_id):
    """Async function to handle Playwright browser automation for case data extraction"""
    async with async_playwright() as p:
        # Launch browser
        logging.info("Launching Playwright browser for case data extraction")
        browser = await p.chromium.launch(
            headless=os.environ.get('JAMI_HIDE_BROWSER', 'False').lower() == 'true'
        )
        
        try:
            # Create a new browser context
            context = await browser.new_context(
                viewport={"width": 1280, "height": 800},
                user_agent="Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/136.0.0.0 Safari/537.36"
            )
            
            # Create a new page
            page = await context.new_page()
            
            # Check if we have cookies from previous login
            cookies = session.get('cookies')
            login_needed = True
            
            if cookies:
                try:
                    # Try to set cookies and go directly to case page
                    logging.info("Attempting to use existing session cookies")
                    await context.add_cookies(cookies)
                    
                    # Go directly to case page
                    case_url = app_config['rdn']['case_url_template'].format(case_id=case_id)
                    logging.info(f"Navigating directly to case URL: {case_url}")
                    await page.goto(case_url)
                    
                    # Wait briefly and check if we got redirected to login page
                    try:
                        await page.wait_for_load_state("networkidle", timeout=10000)  # Increased timeout
                        logging.info("Session cookie navigation successful")
                    except Exception as e:
                        logging.warning(f"Timeout waiting for networkidle with cookies, continuing anyway: {str(e)}")
                        # Wait a reasonable time anyway
                        await page.wait_for_timeout(3000)
                    current_url = page.url
                    
                    if "login" not in current_url.lower():
                        # Successfully used existing session
                        login_needed = False
                        logging.info("Successfully used existing session cookies")
                        await page.screenshot(path=os.path.join('debug', f'direct_access_case_{case_id}.png'))
                    else:
                        logging.info("Redirected to login page, need to log in again")
                except Exception as e:
                    logging.warning(f"Error using existing cookies, will perform new login: {str(e)}")
            
            # If we need to login first
            if login_needed:
                login_url = app_config['rdn']['login_url']
                logging.info(f"Navigating to login URL: {login_url}")
                await page.goto(login_url)
                
                # Save screenshot before filling
                await page.screenshot(path=os.path.join('debug', 'case_before_fill.png'))
                
                # Fill form fields using optimized approach
                logging.info("Filling login form with optimized approach")
                
                # Define credentials from session
                credentials = {
                    "username": session.get('username'),
                    "password": session.get('password'),
                    "securityCode": session.get('security_code')
                }
                
                # Use common selectors for login forms - structured for performance
                field_selectors = {
                    "username": ["#username", "[name=username]", "input[type=text]:nth-of-type(1)", "input:not([type=hidden]):nth-of-type(1)"],
                    "password": ["#password", "[name=password]", "input[type=password]", "input:not([type=hidden]):nth-of-type(2)"],
                    "securityCode": ["#security_code", "#securityCode", "#code", "[name=security_code]", "[name=securityCode]", "[name=code]", "input:not([type=hidden]):nth-of-type(3)"]
                }
                
                security_code_filled = False
                
                # Try to fill all fields with minimal selector queries
                for field, value in credentials.items():
                    if not value:  # Skip empty values
                        continue
                        
                    field_filled = False
                    # Try each selector with minimal timeout
                    for selector in field_selectors.get(field, []):
                        try:
                            element = await page.wait_for_selector(selector, timeout=50)
                            if element:
                                await element.fill(value)
                                logging.info(f"{field} filled using {selector}")
                                field_filled = True
                                if field == "securityCode":
                                    security_code_filled = True
                                break
                        except Exception:
                            continue
                            
                    # If field wasn't filled, try direct typing as last resort
                    if not field_filled:
                        try:
                            # Use position-based fallback
                            field_index = 0  # Default to first field
                            if field == "password":
                                field_index = 1
                            elif field == "securityCode":
                                field_index = 2
                                
                            # Get visible inputs and use indexed position
                            inputs = await page.query_selector_all("input:not([type=hidden])")
                            if len(inputs) > field_index:
                                await inputs[field_index].click()
                                await page.keyboard.type(value)
                                logging.info(f"{field} typed into input #{field_index}")
                                if field == "securityCode":
                                    security_code_filled = True
                        except Exception as e:
                            logging.error(f"Failed to fill {field}: {str(e)}")
                
                if not security_code_filled:
                    logging.warning("Security code field may not have been properly filled")
                
                # Take screenshot after filling the form
                await page.screenshot(path=os.path.join('debug', 'case_after_fill.png'))
                
                # Check for CAPTCHA and handle login with optimized approach
                logging.info("Checking for CAPTCHA before clicking login button")
                
                # Check for CAPTCHA and handle it
                try:
                    captcha_frame = await page.query_selector("iframe[src*='recaptcha'], iframe[title*='reCAPTCHA'], div.g-recaptcha")
                    if captcha_frame:
                        logging.info("CAPTCHA detected during case extraction! Notifying user")
                        html_content = await page.content()
                        with open(os.path.join('debug', 'case_captcha_page.html'), 'w', encoding='utf-8') as f:
                            f.write(html_content)
                        await page.screenshot(path=os.path.join('debug', 'case_captcha_detected.png'))
                        
                        # If headless mode is on, we can't solve CAPTCHA
                        if os.environ.get('JAMI_HIDE_BROWSER', 'False').lower() == 'true':
                            await browser.close()
                            return False, "CAPTCHA detected during case extraction - manual login required. Please set JAMI_HIDE_BROWSER=False to show the browser and solve the CAPTCHA."
                        else:
                            # Wait for user to solve CAPTCHA manually
                            logging.info("Browser is visible - waiting 60 seconds for user to solve CAPTCHA manually...")
                            await page.wait_for_timeout(60000)  # Wait 60 seconds for user to solve
                            logging.info("Continuing after CAPTCHA timeout")
                except Exception as e:
                    logging.warning(f"Error checking for CAPTCHA during case extraction: {str(e)}")
                
                # Try to submit form directly first
                try:
                    logging.info("Attempting to submit form using JavaScript")
                    await page.evaluate("document.querySelector('form').submit()")
                    logging.info("Form submitted via JavaScript")
                    button_clicked = True
                except Exception as e:
                    logging.warning(f"JavaScript form submission failed in case extraction: {str(e)}")
                    
                    # Fall back to clicking buttons
                    logging.info("Trying button click approach")
                    # Try multiple button selectors in order of likelihood
                    button_selectors = [
                        "button.btn-success",  # The actual Login button class
                        "button:has-text('Login')", 
                        "button[type='submit']", 
                        "input[type='submit']",
                        "button:has-text('Sign In')",
                        "button:has-text('Submit')",
                        "input[type='button'][value='Login']",
                        "form button"
                    ]
                    
                    button_clicked = False
                    for selector in button_selectors:
                        try:
                            button = await page.wait_for_selector(selector, timeout=50)
                            if button:
                                await button.click()
                                logging.info(f"Clicked button using selector: {selector}")
                                button_clicked = True
                                break
                        except Exception:
                            continue
                
                # If no button clicked, try pressing Enter as fallback
                if not button_clicked:
                    try:
                        inputs = await page.query_selector_all("input:not([type=hidden])")
                        if inputs:
                            await inputs[-1].press("Enter")
                            logging.info("Pressed Enter on last input field")
                        else:
                            logging.warning("No input fields found to press Enter")
                    except Exception as e:
                        logging.error(f"Failed to submit form: {str(e)}")
                
                # Screenshot after button click
                await page.screenshot(path=os.path.join('debug', 'case_after_button_click.png'))
                
                # Wait for navigation using more flexible approach
                try:
                    # Try waiting for any successful navigation away from login
                    logging.info("Waiting for navigation after login in case extraction...")
                    
                    # Just wait for load state instead of specific URL
                    await page.wait_for_load_state("domcontentloaded", timeout=10000)
                    
                    # Check if we're still on the login page
                    current_url = page.url
                    if "login" in current_url.lower():
                        logging.warning(f"Still on login page after button click: {current_url}")
                    else:
                        logging.info(f"Navigation successful: {current_url}")
                except Exception as e:
                    logging.warning(f"Navigation wait timed out, but proceeding: {str(e)}")
            
            # Now navigate to case page (either after login or directly with cookies)
            if not login_needed or (login_needed and "login" not in page.url.lower()):
                # Navigate to case page using URL from config (if not already there)
                if not page.url or case_id not in page.url:
                    case_url = app_config['rdn']['case_url_template'].format(case_id=case_id)
                    logging.info(f"Navigating to case URL: {case_url}")
                    await page.goto(case_url)
                    
                    # Wait for case page to load and take screenshot - extended timeout
                    try:
                        await page.wait_for_load_state("networkidle", timeout=15000)  # Increased to 15 seconds
                        logging.info("Case page loaded successfully with networkidle state")
                    except Exception as e:
                        logging.warning(f"Timeout waiting for networkidle, but continuing: {str(e)}")
                        # Still wait a few seconds to let the page load somewhat
                        await page.wait_for_timeout(3000)
                
                # Take a screenshot of the case page
                await page.screenshot(path=os.path.join('debug', f'case_{case_id}.png'))
            else:
                logging.error("Login failed or session expired, unable to access case page")
                await browser.close()
                return False, "Login failed or session expired, unable to access case page"
            
            # No matter what happens with case extraction, we'll attempt to get some data
            try:
                # Extract page content for parsing
                page_content = await page.content()
                
                # Save page source for debugging
                with open(os.path.join('debug', f'case_{case_id}.html'), 'w', encoding='utf-8') as f:
                    f.write(page_content)
                
                soup = BeautifulSoup(page_content, 'html.parser')
                
                # Initialize case data
                case_data = {
                    "caseId": case_id,
                    "clientName": "Not Found",
                    "lienHolder": "Not Found",
                    "orderTo": "Not Found",
                    "fees": []
                }
                
                # Take a screenshot of the current state
                await page.screenshot(path=os.path.join('debug', f'case_extraction_start_{case_id}.png'))
                
                logging.info("Beginning case data extraction - this will proceed even with partial data")
            except Exception as e:
                logging.error(f"Critical error at case extraction setup: {str(e)}")
                # Still try to continue with a minimal case data structure
                case_data = {
                    "caseId": case_id,
                    "clientName": "Error extracting data",
                    "lienHolder": "Error extracting data",
                    "orderTo": "Error extracting data",
                    "fees": []
                }
            
            # Check if we already have a definitive client name from a previous extraction
            if 'definitive_client_name' in session and session['definitive_client_name']:
                logging.info(f"Using definitive client name from session: {session['definitive_client_name']}")
                case_data["clientName"] = session['definitive_client_name']
            
            # Extract client information using various selectors and patterns following server-upgradedv2.py
            logging.info("Extracting case information using multiple approaches")
            
            # APPROACH 1: Look for definition list (dt/dd) structure
            logging.info("Attempting to extract case information using dt/dd approach")
            
            # Find all dt elements
            dt_elements = soup.find_all('dt')
            for dt in dt_elements:
                dt_text = dt.get_text().strip()
                
                # Find client information - checking for exact match to avoid client acct no
                if dt_text.lower() == 'client':
                    # Find the sibling dd element
                    try:
                        # Try to get the immediately following dd element
                        next_element = dt.next_sibling
                        while next_element and next_element.name != 'dd':
                            next_element = next_element.next_sibling
                        
                        if next_element and next_element.name == 'dd':
                            text = next_element.get_text().strip()
                            if text and not text.startswith("$"):
                                case_data["clientName"] = text
                                logging.info(f"Found client name using dt/dd next sibling: {text}")
                                session['definitive_client_name'] = text
                                continue
                        
                        # If not found via next_sibling, try parent method
                        parent = dt.parent
                        if parent:
                            dd_elements = parent.find_all('dd')
                            if dd_elements and dd_elements[0].get_text().strip():
                                text = dd_elements[0].get_text().strip()
                                if text and not text.startswith("$"):
                                    case_data["clientName"] = text
                                    logging.info(f"Found client name using dt/dd parent method: {text}")
                                    session['definitive_client_name'] = text
                    except Exception as e:
                        logging.error(f"Error finding client dd element: {str(e)}")
                
                # Find lien holder information
                elif 'lien holder' in dt_text.lower() or 'lienholder' in dt_text.lower():
                    try:
                        # Try to get the immediately following dd element
                        next_element = dt.next_sibling
                        while next_element and next_element.name != 'dd':
                            next_element = next_element.next_sibling
                        
                        if next_element and next_element.name == 'dd':
                            text = next_element.get_text().strip()
                            if text:
                                case_data["lienHolder"] = text
                                logging.info(f"Found lien holder name using dt/dd next sibling: {text}")
                                continue
                        
                        # If not found via next_sibling, try parent method
                        parent = dt.parent
                        if parent:
                            dd_elements = parent.find_all('dd')
                            if dd_elements and dd_elements[0].get_text().strip():
                                case_data["lienHolder"] = dd_elements[0].get_text().strip()
                                logging.info(f"Found lien holder name using dt/dd parent method: {case_data['lienHolder']}")
                    except Exception as e:
                        logging.error(f"Error finding lien holder dd element: {str(e)}")
            
            # APPROACH 2: Look for div.col-auto pattern as in the example HTML
            if case_data["clientName"] == "Not Found" or case_data["lienHolder"] == "Not Found":
                logging.info("Attempting col-auto/div pattern for client and lien holder")
                
                # Look for client - using exact match to avoid "Client Acct No"
                client_divs = soup.find_all('div', class_=lambda c: c and 'col-auto' in c)
                for div in client_divs:
                    dt_elements = div.find_all('dt')
                    for dt in dt_elements:
                        if dt.get_text().strip().lower() == 'client':
                            # Directly get the dd that is a direct child of this div
                            dd = div.find('dd', recursive=False)
                            if not dd:  # If not a direct child, get any dd
                                dd = div.find('dd')
                            
                            if dd and dd.get_text().strip():
                                text = dd.get_text().strip()
                                if text and not text.startswith("$"):
                                    case_data["clientName"] = text
                                    logging.info(f"Found client name using col-auto pattern: {text}")
                                    session['definitive_client_name'] = text
                                    break
                
                # Look for lien holder
                for div in client_divs:
                    dt_elements = div.find_all('dt')
                    for dt in dt_elements:
                        if dt.get_text().strip().lower() == 'lien holder':
                            # Directly get the dd that is a direct child of this div
                            dd = div.find('dd', recursive=False)
                            if not dd:  # If not a direct child, get any dd
                                dd = div.find('dd')
                            
                            if dd and dd.get_text().strip():
                                text = dd.get_text().strip()
                                case_data["lienHolder"] = text
                                logging.info(f"Found lien holder name using col-auto pattern: {text}")
                                break
                                
            # APPROACH 2b: Look for common layouts like tables and headers
            if case_data["clientName"] == "Not Found" or case_data["lienHolder"] == "Not Found":
                logging.info("Attempting table/field layout patterns")
                
                # Look for table headers and values
                headers = soup.find_all(['th', 'td', 'strong', 'b', 'label'])
                for header in headers:
                    header_text = header.get_text().strip().lower()
                    
                    # Check for client
                    if case_data["clientName"] == "Not Found" and 'client' in header_text:
                        # Try to get the next cell (sibling)
                        next_cell = header.find_next_sibling()
                        if next_cell and next_cell.get_text().strip():
                            case_data["clientName"] = next_cell.get_text().strip()
                            logging.info(f"Found client name from table cell: {case_data['clientName']}")
                            
                        # Try to get parent's next sibling
                        elif header.parent and header.parent.find_next_sibling():
                            sibling = header.parent.find_next_sibling()
                            if sibling and sibling.get_text().strip():
                                case_data["clientName"] = sibling.get_text().strip()
                                logging.info(f"Found client name from parent's sibling: {case_data['clientName']}")
                    
                    # Check for lien holder
                    if case_data["lienHolder"] == "Not Found" and ('lien' in header_text or 'holder' in header_text):
                        # Try to get the next cell (sibling)
                        next_cell = header.find_next_sibling()
                        if next_cell and next_cell.get_text().strip():
                            case_data["lienHolder"] = next_cell.get_text().strip()
                            logging.info(f"Found lien holder from table cell: {case_data['lienHolder']}")
                            
                        # Try to get parent's next sibling
                        elif header.parent and header.parent.find_next_sibling():
                            sibling = header.parent.find_next_sibling()
                            if sibling and sibling.get_text().strip():
                                case_data["lienHolder"] = sibling.get_text().strip()
                                logging.info(f"Found lien holder from parent's sibling: {case_data['lienHolder']}")
                
                # Look for specific headings and then get values from surrounding elements
                try:
                    # Looking for specific layout patterns in the webpage
                    client_elements = soup.find_all(lambda tag: tag.name in ['td', 'div', 'span'] and 
                                                   tag.get_text().strip().lower() == 'client')
                    for element in client_elements:
                        # Look at next siblings or parent->next sibling
                        if element.next_sibling:
                            text = element.next_sibling.strip() if isinstance(element.next_sibling, str) else element.next_sibling.get_text().strip()
                            if text:
                                case_data["clientName"] = text
                                logging.info(f"Found client name from direct sibling: {text}")
                                break
                        
                        # Try looking at parent row -> next column
                        parent_row = element.find_parent('tr')
                        if parent_row:
                            next_cell = element.find_next_sibling('td')
                            if next_cell and next_cell.get_text().strip():
                                case_data["clientName"] = next_cell.get_text().strip()
                                logging.info(f"Found client name from next table cell: {case_data['clientName']}")
                                break
                except Exception as e:
                    logging.error(f"Error in additional client extraction: {str(e)}")
            
            # Get Order To info for repo type - similar to client/lienholder extraction but more extensive
            # Based on the screenshot, "Order To" appears with the value "Involuntary Repo" in a green button next to it
            
            # First try to find any element with "Order To" text
            logging.info("Attempting multiple strategies to extract Order To field")
            order_to_containers = []
            
            # Method 1: Find by exact labels
            try:
                order_to_divs = soup.find_all(['dt', 'div', 'span', 'label', 'th', 'td'], 
                                              string=lambda s: s and re.search(r'\bOrder\s*To\b', s, re.IGNORECASE))
                order_to_containers.extend(order_to_divs)
                logging.info(f"Found {len(order_to_divs)} elements with 'Order To' text")
            except Exception as e:
                logging.warning(f"Error in Order To search: {str(e)}")
            
            # Method 2: Find by proximity to the "Involuntary Repo" button - based on your screenshot
            try:
                repo_buttons = soup.find_all(['button', 'div', 'span', 'a'], 
                                             string=lambda s: s and ('involuntary repo' in s.lower() or 'voluntary repo' in s.lower()))
                for button in repo_buttons:
                    # Find nearby Order To text
                    parent_div = button.find_parent('div', class_=lambda c: c is not None)
                    if parent_div:
                        order_labels = parent_div.find_all(string=lambda s: s and 'order to' in s.lower())
                        if order_labels:
                            for label in order_labels:
                                parent_element = label.parent
                                if parent_element not in order_to_containers:
                                    order_to_containers.append(parent_element)
                                    logging.info(f"Found Order To label near repo button in {parent_element.name}")
            except Exception as e:
                logging.warning(f"Error in repo button proximity search: {str(e)}")
                
            # Method 3: Look for any table structure with Order To
            try:
                # Find tables or table-like structures
                tables = soup.find_all(['table', 'div', 'section'], class_=lambda c: c and ('table' in str(c).lower() or 'grid' in str(c).lower() or 'info' in str(c).lower()))
                for table in tables:
                    # Look for Order To headers
                    headers = table.find_all(string=lambda s: s and 'order to' in s.lower())
                    for header in headers:
                        parent = header.parent
                        if parent not in order_to_containers:
                            order_to_containers.append(parent)
                            logging.info(f"Found Order To label in table/grid: {parent.name}")
            except Exception as e:
                logging.warning(f"Error in table structure search: {str(e)}")
                
            # Process all found Order To containers
            found_order_to = False
            for element in order_to_containers:
                if found_order_to:
                    break
                    
                try:
                    # Try multiple approaches to find the value
                    
                    # Approach 1: Look for sibling with repo info
                    siblings = list(element.next_siblings)
                    for sibling in siblings[:3]:  # Check first 3 siblings
                        if hasattr(sibling, 'get_text'):
                            text = sibling.get_text().strip()
                            if text and ('repo' in text.lower() or 'voluntary' in text.lower()):
                                case_data["orderTo"] = text
                                logging.info(f"Found order to from sibling: {text}")
                                found_order_to = True
                                break
                    
                    if not found_order_to:
                        # Approach 2: Check for dd elements (dt/dd pattern)
                        parent = element.parent
                        if parent:
                            dd_elements = parent.find_all('dd')
                            if dd_elements and dd_elements[0].get_text().strip():
                                text = dd_elements[0].get_text().strip()
                                if text:
                                    case_data["orderTo"] = text
                                    logging.info(f"Found order to using dt/dd: {text}")
                                    found_order_to = True
                    
                    if not found_order_to:
                        # Approach 3: Look for any button/badge near this element
                        parent_container = element.find_parent(['div', 'section', 'tr', 'td'])
                        if parent_container:
                            badges = parent_container.find_all(['button', 'span', 'div'], 
                                                               class_=lambda c: c and ('badge' in str(c).lower() or 'label' in str(c).lower() or 'btn' in str(c).lower()))
                            for badge in badges:
                                text = badge.get_text().strip()
                                if text and ('repo' in text.lower() or 'voluntary' in text.lower()):
                                    case_data["orderTo"] = text
                                    logging.info(f"Found order to from badge/button: {text}")
                                    found_order_to = True
                                    break
                except Exception as e:
                    logging.warning(f"Error extracting Order To value: {str(e)}")
            
            # Direct lookup for specific structure seen in screenshot
            if case_data["orderTo"] == "Not Found":
                try:
                    # Find elements with label-like class names containing 'Order To'
                    order_rows = soup.find_all(['tr', 'div'], class_=lambda c: c and 'row' in str(c).lower())
                    for row in order_rows:
                        label_cells = row.find_all(lambda tag: tag.name in ['td', 'div', 'span'] and 'order to' in tag.get_text().lower())
                        if label_cells:
                            # Look for adjacent cells/elements with repo info
                            for label_cell in label_cells:
                                adjacent_cells = list(label_cell.next_siblings) if label_cell.next_siblings else []
                                for cell in adjacent_cells:
                                    if hasattr(cell, 'get_text'):
                                        text = cell.get_text().strip()
                                        if 'repo' in text.lower() or 'voluntary' in text.lower():
                                            case_data["orderTo"] = text
                                            logging.info(f"Found Order To from table row: {text}")
                                            found_order_to = True
                                            break
                                
                                # Also look for any green button nearby
                                if not found_order_to:
                                    parent_row = label_cell.parent
                                    if parent_row:
                                        buttons = parent_row.find_all(['button', 'span', 'div'], class_=lambda c: c and any(x in str(c).lower() for x in ['badge', 'btn', 'green', 'label', 'tag']))
                                        for button in buttons:
                                            text = button.get_text().strip()
                                            if 'repo' in text.lower() or 'voluntary' in text.lower():
                                                case_data["orderTo"] = text
                                                logging.info(f"Found Order To from green button: {text}")
                                                found_order_to = True
                                                break
                except Exception as e:
                    logging.warning(f"Error in direct table structure search: {str(e)}")
                    
            # Direct search for the green Involuntary Repo button as shown in screenshot
            if case_data["orderTo"] == "Not Found":
                try:
                    green_buttons = soup.find_all(['button', 'span', 'div', 'a'], 
                                                 class_=lambda c: c and any(x in str(c).lower() for x in ['badge', 'btn', 'success', 'green', 'primary']))
                    for button in green_buttons:
                        text = button.get_text().strip()
                        if ('involuntary' in text.lower() or 'voluntary' in text.lower()) and 'repo' in text.lower():
                            # This is likely the button we want - check if it's near "Order To"
                            parent_container = button.find_parent(['div', 'tr', 'section'])
                            if parent_container:
                                order_labels = parent_container.find_all(string=lambda s: s and 'order to' in s.lower())
                                if order_labels:
                                    # This is highly likely to be the correct value
                                    case_data["orderTo"] = text
                                    logging.info(f"Found Order To from green button with nearby 'Order To' text: {text}")
                                    break
                                else:
                                    # Still might be correct - just no nearby "Order To" label
                                    case_data["orderTo"] = text
                                    logging.info(f"Found likely Order To from green repo button: {text}")
                                    break
                except Exception as e:
                    logging.warning(f"Error in green button search: {str(e)}")
                        
            # APPROACH 3: Look for badge elements that indicate repo type (directly from server-upgradedv2.py)
            logging.info("Attempting to extract repo type using badge elements and Order To field")
            
            # From the badge-invol and badge-vol patterns in server-upgradedv2.py
            badge_elements = soup.find_all(['span', 'div', 'button'], class_=lambda c: c and ('badge' in str(c).lower() or 'label' in str(c).lower()))
            for badge in badge_elements:
                badge_text = badge.get_text().strip().lower()
                if 'involuntary' in badge_text:
                    case_data["repoType"] = "Involuntary Repo"
                    logging.info(f"Found Involuntary Repo using badge: {badge.name}.{badge.get('class', [])}")
                    break
                elif 'voluntary' in badge_text:
                    case_data["repoType"] = "Voluntary Repo"
                    logging.info(f"Found Voluntary Repo using badge: {badge.name}.{badge.get('class', [])}")
                    break
            
            # Order To field extraction (most important - as shown in your screenshot)
            # In your case, the Order To section contains the repo type
            if case_data.get("repoType") == "Not Found":
                # First look for elements with id="case_order_type_static" (from server-upgradedv2.py)
                order_type_elements = soup.find_all(['span', 'div'], id="case_order_type_static")
                if order_type_elements:
                    for element in order_type_elements:
                        text = element.get_text().strip().lower()
                        if 'involuntary' in text:
                            case_data["repoType"] = "Involuntary Repo"
                            logging.info(f"Found Involuntary Repo in order type static element")
                            break
                        elif 'voluntary' in text:
                            case_data["repoType"] = "Voluntary Repo"
                            logging.info(f"Found Voluntary Repo in order type static element")
                            break
                
                # If still not found, look in the Order To area
                if case_data.get("repoType") == "Not Found" and case_data.get("orderTo") not in [None, "", "Not Found"]:
                    order_text = case_data.get("orderTo").lower()
                    if 'involuntary' in order_text:
                        case_data["repoType"] = "Involuntary Repo"
                        logging.info(f"Found Involuntary Repo from Order To value")
                    elif 'voluntary' in order_text:
                        case_data["repoType"] = "Voluntary Repo"
                        logging.info(f"Found Voluntary Repo from Order To value")
                    elif 'repo' in order_text:
                        # If it just says 'repo' assume it's involuntary (most common)
                        case_data["repoType"] = "Involuntary Repo"
                        logging.info(f"Found likely Involuntary Repo from Order To value containing 'repo'")
            
            # Check for green buttons specifically near Order To section
            if case_data.get("repoType") == "Not Found":
                # Use the image you provided to look for the green button specifically
                green_elements = soup.find_all(['button', 'div', 'span'], 
                                              class_=lambda c: c and any(x in str(c).lower() for x in ['badge', 'btn', 'label', 'tag']))
                for element in green_elements:
                    text = element.get_text().strip().lower()
                    if 'involuntary' in text and 'repo' in text:
                        case_data["repoType"] = "Involuntary Repo"
                        logging.info(f"Found Involuntary Repo in button/badge element")
                        break
                    elif 'voluntary' in text and 'repo' in text:
                        case_data["repoType"] = "Voluntary Repo"
                        logging.info(f"Found Voluntary Repo in button/badge element")
                        break
            
            # APPROACH 4: Fallback to regex patterns on entire page text
            if case_data.get("clientName") == "Not Found" or case_data.get("lienHolder") == "Not Found" or case_data.get("repoType") == "Not Found":
                logging.info("Falling back to text pattern matching approach")
                all_text = soup.get_text()
                
                # Client patterns from server-upgradedv2.py
                if case_data.get("clientName") == "Not Found":
                    client_patterns = [
                        r'Client\s*:\s*([^\n:]+)',
                        r'Client\s+([A-Za-z0-9\s\.\,\&\;\-\'\"-]+)(?=\s*Collector|\s*Lien|\s*$)',
                        r'Client(?:[\s\:]*)(.*?)(?=\s*Collector|\s*Lien|\s*$)'
                    ]
                    
                    for pattern in client_patterns:
                        client_matches = re.search(pattern, all_text, re.IGNORECASE)
                        if client_matches and client_matches.group(1):
                            case_data["clientName"] = client_matches.group(1).strip()
                            logging.info(f"Found client name using regex: {case_data['clientName']}")
                            break
                
                # Lien holder patterns from server-upgradedv2.py
                if case_data.get("lienHolder") == "Not Found":
                    lien_patterns = [
                        r'Lien\s*Holder\s*:\s*([^\n:]+)',
                        r'Lien\s*Holder\s+([A-Za-z0-9\s\.\,\&\;\-\'\"-]+)(?=\s*Client|\s*Acct|\s*File|\s*$)',
                        r'Lien\s*Holder(?:[\s\:]*)(.*?)(?=\s*Client|\s*Acct|\s*File|\s*$)'
                    ]
                    
                    for pattern in lien_patterns:
                        lien_holder_matches = re.search(pattern, all_text, re.IGNORECASE)
                        if lien_holder_matches and lien_holder_matches.group(1):
                            case_data["lienHolder"] = lien_holder_matches.group(1).strip()
                            logging.info(f"Found lien holder name using regex: {case_data['lienHolder']}")
                            break
                
                # Repo type determination based on page text
                if case_data.get("repoType") == "Not Found":
                    if "involuntary" in all_text.lower():
                        case_data["repoType"] = "Involuntary Repo"
                        logging.info("Determined repo type as Involuntary from page text")
                    elif "voluntary" in all_text.lower():
                        case_data["repoType"] = "Voluntary Repo"
                        logging.info("Determined repo type as Voluntary from page text")
                    else:
                        # Default to Involuntary as more common
                        case_data["repoType"] = "Involuntary Repo"
                        logging.info("Defaulting to Involuntary Repo type")
                
            # Clean up values following server-upgradedv2.py pattern
            if case_data["clientName"] != "Not Found":
                case_data["clientName"] = re.sub(r'^Client\s*:?\s*', '', case_data["clientName"]).strip()
                
                # Fix for "$0.0" appearing as client name
                if case_data["clientName"].startswith("$") and any(x in case_data["clientName"] for x in ["0.0", "0.00"]):
                    logging.warning(f"Found probable incorrect client name: {case_data['clientName']}")
                    case_data["clientName"] = "Not Found"  # Reset to try other methods
                
            if case_data["lienHolder"] != "Not Found":
                case_data["lienHolder"] = re.sub(r'^Lien\s*Holder\s*:?\s*', '', case_data["lienHolder"]).strip()
                
            # Extract client name based on the exact HTML structure provided
            if case_data["clientName"] == "Not Found" or case_data["clientName"] == "Default Client":
                try:
                    logging.info("Attempting to extract client name using exact structure provided in HTML")
                    
                    # EXACT MATCH for the structure you provided:
                    # <div class="col-auto"><dt>Client</dt><dd>Primeritus- IBEAM</dd></div>
                    client_divs = soup.find_all('div', class_='col-auto')
                    
                    # Use a flag to ensure we only capture the actual client name, not other fields
                    client_found = False
                    
                    for div in client_divs:
                        dt = div.find('dt')
                        if dt and dt.get_text().strip() == 'Client':
                            dd = div.find('dd')
                            if dd:
                                text = dd.get_text().strip()
                                if text and not text.startswith("$") and text != "<empty>":
                                    # Check if this looks like an actual client name
                                    # Client names are usually more than 2 characters and don't contain only numbers
                                    if len(text) > 2 and not text.isdigit() and 'Not' not in text:
                                        case_data["clientName"] = text
                                        logging.info(f"Found client name from exact structure match: {text}")
                                        client_found = True
                                        break
                    
                    # If client name found, set a flag to avoid overriding with incorrect values
                    if client_found:
                        # Save this as the definitive client name
                        session['definitive_client_name'] = case_data["clientName"]
                                    
                    # If still not found, try a more flexible approach but targeting the same structure
                    if case_data["clientName"] == "Not Found" or case_data["clientName"] == "Default Client":
                        # Look for a section div that contains rows of data
                        sections = soup.find_all('div', class_=lambda c: c and ('section' in str(c).lower() or 'main' in str(c).lower()))
                        for section in sections:
                            rows = section.find_all('div', class_=lambda c: c and ('row' in str(c).lower() or 'justify' in str(c).lower()))
                            for row in rows:
                                col_divs = row.find_all('div', class_='col-auto')
                                for col in col_divs:
                                    dt = col.find('dt')
                                    if dt and 'client' in dt.get_text().lower() and 'acct' not in dt.get_text().lower():
                                        dd = col.find('dd')
                                        if dd:
                                            text = dd.get_text().strip()
                                            if text and not text.startswith("$"):
                                                case_data["clientName"] = text
                                                logging.info(f"Found client name from section/row structure: {text}")
                                                break
                                                
                    # Direct XPath-like approach
                    if case_data["clientName"] == "Not Found" or case_data["clientName"] == "Default Client":
                        # Find any dt that contains exactly "Client" text
                        client_dts = soup.find_all('dt', string=lambda s: s and s.strip() == 'Client')
                        for dt in client_dts:
                            # Get the parent div.col-auto
                            parent_div = dt.parent
                            if parent_div and parent_div.name == 'div' and 'col-auto' in parent_div.get('class', []):
                                # Find the dd within this div
                                dd = parent_div.find('dd')
                                if dd:
                                    text = dd.get_text().strip()
                                    if text and not text.startswith("$"):
                                        case_data["clientName"] = text
                                        logging.info(f"Found client name from direct dt-dd match: {text}")
                                        break
                                        
                    # As a last resort, parse the HTML string directly
                    if case_data["clientName"] == "Not Found" or case_data["clientName"] == "Default Client":
                        # Use regex to extract from the pattern "<dt>Client</dt><dd>VALUE_WE_WANT</dd>"
                        html_content = str(soup)
                        # Fixed regex to correctly capture the client name without attributes
                        client_match = re.search(r'<dt>Client</dt>\s*<dd[^>]*>(.*?)</dd>', html_content)
                        if client_match:
                            text = client_match.group(1).strip()
                            if text and not text.startswith("$"):
                                case_data["clientName"] = text
                                logging.info(f"Found client name using direct HTML regex: {text}")
                    
                    # Ultra-fallback: Search for anything that looks like a client name in the page
                    if case_data["clientName"] == "Not Found" or case_data["clientName"] == "Default Client":
                        logging.info("Trying ultra-fallback client name extraction")
                        # Look for any text that might contain "Client:" followed by a name
                        all_text_content = soup.get_text()
                        
                        # Try different patterns
                        patterns = [
                            r'Client\s*:\s*([A-Za-z0-9\s\-]+(?:LLC|Inc|Corp)?)',  # Client: Name
                            r'Client\s*Name\s*:\s*([A-Za-z0-9\s\-]+(?:LLC|Inc|Corp)?)',  # Client Name: Name
                            r'Client\s*:\s*([^$\n\r]{3,30})',  # Client: (anything not starting with $ and reasonable length)
                            r'Client\s*Name\s*:\s*([^$\n\r]{3,30})'  # Client Name: (anything not starting with $ and reasonable length)
                        ]
                        
                        for pattern in patterns:
                            match = re.search(pattern, all_text_content)
                            if match:
                                candidate = match.group(1).strip()
                                # Validate candidate
                                if (candidate and len(candidate) > 2 and 
                                    not candidate.startswith("$") and 
                                    not candidate.isdigit() and 
                                    'not' not in candidate.lower()):
                                    case_data["clientName"] = candidate
                                    logging.info(f"Found client name using text pattern: {candidate}")
                                    break
                                    
                        # Try finding known client names if still not found
                        known_clients = ["Primeritus", "IBEAM", "MasterTrak", "CarsArrive", "PAR North America"]
                        for client in known_clients:
                            if client.lower() in all_text_content.lower():
                                case_data["clientName"] = client
                                logging.info(f"Found client name using known client list: {client}")
                                break
                                
                except Exception as e:
                    logging.error(f"Error in exact structure client name extraction: {str(e)}")
                    
            # If we somehow ended up with a dollar amount as client name, clear it
            if case_data["clientName"] != "Not Found" and case_data["clientName"].startswith("$") and re.search(r'^\$\d+(\.\d+)?$', case_data["clientName"]):
                logging.warning(f"Clearing dollar amount incorrectly detected as client name: {case_data['clientName']}")
                case_data["clientName"] = "Default Client"  # Use a default instead of "Not Found"
                
            # CRITICAL: Ensure order to value is also used as repo type (based on the screenshot)
            # This is the key relationship - the "Order To" field contains the repo type
            if case_data["orderTo"] != "Not Found" and ("repo" in case_data["orderTo"].lower() or "voluntary" in case_data["orderTo"].lower()):
                # Use the orderTo value as repoType since they contain the same information
                case_data["repoType"] = case_data["orderTo"]
                logging.info(f"Set repoType from orderTo field: {case_data['repoType']}")
                
            # Also vice versa - if we found repo type but not order to
            if case_data["orderTo"] == "Not Found" and case_data["repoType"] != "Not Found":
                case_data["orderTo"] = case_data["repoType"]
                logging.info(f"Set orderTo from repoType field: {case_data['orderTo']}")
                
            # Final safety check - ensure both have valid values
            if case_data["repoType"] == "Not Found" and case_data["orderTo"] == "Not Found":
                # Default to most common type
                case_data["repoType"] = "Involuntary Repo"
                case_data["orderTo"] = "Involuntary Repo"
                logging.info("Using default 'Involuntary Repo' as last resort fallback")
            
            # Enhanced fee information extraction with more comprehensive analysis
            # Define the dollar pattern for matching
            dollar_pattern = r'\$(\d{1,3}(,\d{3})*(\.\d{2})?)'
            
            # First, analyze the entire page structure for fee data
            fee_sections = []
            
            # Look for fee-related sections or tables
            fee_tables = soup.find_all('table')
            for table in fee_tables:
                # Check if this looks like a fee table
                header_row = table.find('tr')
                if header_row:
                    header_text = header_row.get_text().lower()
                    if any(term in header_text for term in ['fee', 'amount', 'charge', 'cost', 'payment', 'transaction']):
                        fee_sections.append(table)
                        logging.info(f"Found fee table with header: {header_text}")
            
            # Look for fee-related divs
            fee_divs = soup.find_all(['div', 'section'], class_=re.compile(r'fee|charge|cost|payment', re.IGNORECASE))
            fee_sections.extend(fee_divs)
            
            # If we found structured fee sections, extract from them
            structured_fees = []
            if fee_sections:
                logging.info(f"Found {len(fee_sections)} structured fee sections")
                
                for section in fee_sections:
                    # For tables, process rows
                    if section.name == 'table':
                        rows = section.find_all('tr')
                        # Skip header row if it exists
                        start_idx = 1 if (rows and any(th.name == 'th' for th in rows[0].find_all())) else 0
                        
                        for row in rows[start_idx:]:
                            # Get all cells in the row
                            cells = row.find_all(['td', 'th'])
                            if len(cells) >= 2:  # Need at least description and amount
                                row_text = row.get_text().strip()
                                # Look for dollar amount in any cell
                                amount_match = re.search(dollar_pattern, row_text)
                                if amount_match:
                                    amount_str = amount_match.group(0)
                                    amount = float(amount_match.group(1).replace(',', ''))
                                    
                                    # Try to extract description
                                    description = ""
                                    for cell in cells:
                                        cell_text = cell.get_text().strip()
                                        if cell_text and '$' not in cell_text:  # Skip amount cells
                                            description = cell_text
                                            break
                                    
                                    # Get full row text for analysis
                                    # Get fee type and status
                                    fee_type_info = identify_fee_type(row_text)
                                    fee_status = identify_fee_status(row_text)
                                    
                                    # Skip fees with zero amount
                                    if amount > 0:
                                        logging.info(f"Adding structured table fee with amount: ${amount:.2f}")
                                        structured_fees.append({
                                            "description": description or row_text,
                                            "amount": amount,
                                            "amountStr": amount_str,
                                            "category": fee_type_info["category"],
                                            "categoryColor": fee_type_info["color"],
                                            "confidence": fee_type_info["confidence"],
                                            "status": fee_status,
                                            "source": "table",
                                            "raw_text": row_text
                                        })
                    else:
                        # Process divs or sections
                        section_text = section.get_text()
                        
                        # Find all dollar amounts in the section
                        dollar_matches = re.finditer(dollar_pattern, section_text)
                        for match in dollar_matches:
                            amount_str = match.group(0)
                            amount = float(match.group(1).replace(',', ''))
                            
                            # Get surrounding text for context (80 characters for more context)
                            start_pos = max(0, match.start() - 80)
                            end_pos = min(len(section_text), match.end() + 80)
                            surrounding_text = section_text[start_pos:end_pos]
                            
                            # Get fee type and status
                            fee_type_info = identify_fee_type(surrounding_text)
                            fee_status = identify_fee_status(surrounding_text)
                            
                            # Skip fees with zero amount
                            if amount > 0:
                                logging.info(f"Adding structured section fee with amount: ${amount:.2f}")
                                structured_fees.append({
                                    "description": surrounding_text.strip(),
                                    "amount": amount,
                                    "amountStr": amount_str,
                                    "category": fee_type_info["category"],
                                    "categoryColor": fee_type_info["color"],
                                    "confidence": fee_type_info["confidence"],
                                    "status": fee_status,
                                    "source": "section",
                                    "raw_text": surrounding_text
                                })
            
            # As a fallback, scan the entire page text for dollar amounts
            all_text = soup.get_text()
            
            # Find all dollar amounts in the page
            dollar_matches = re.finditer(dollar_pattern, all_text)
            for match in dollar_matches:
                amount_str = match.group(0)  # Get the full dollar amount with $ sign
                amount = float(match.group(1).replace(',', ''))  # Get the numeric value
                
                # Get surrounding text for context (80 characters before and after for more context)
                start_pos = max(0, match.start() - 80)
                end_pos = min(len(all_text), match.end() + 80)
                surrounding_text = all_text[start_pos:end_pos]
                
                # Skip this match if it's too similar to one we already have from structured extraction
                skip = False
                for existing_fee in structured_fees:
                    if abs(existing_fee["amount"] - amount) < 0.01:  # Same amount
                        if surrounding_text in existing_fee["raw_text"] or existing_fee["raw_text"] in surrounding_text:
                            skip = True
                            break
                
                if skip:
                    continue
                
                # Identify fee type and status based on surrounding text
                fee_type_info = identify_fee_type(surrounding_text)
                fee_status = identify_fee_status(surrounding_text)
                
                # Add to the fees list only if it looks like a real fee (not just random dollar amount)
                # Use additional context to filter out non-fee dollar amounts
                is_likely_fee = False
                fee_keywords = ['fee', 'charge', 'cost', 'invoice', 'bill', 'payment', 'paid', 'due', 'storage', 'tow', 'repo']
                if any(keyword in surrounding_text.lower() for keyword in fee_keywords):
                    is_likely_fee = True
                
                if is_likely_fee and amount > 0:  # Only add fees with non-zero amounts
                    logging.info(f"Adding fee with amount: ${amount:.2f}")
                    case_data["fees"].append({
                        "description": surrounding_text.strip(),
                        "amount": amount,
                        "amountStr": amount_str,
                        "category": fee_type_info["category"],
                        "categoryColor": fee_type_info["color"],
                        "confidence": fee_type_info["confidence"],
                        "status": fee_status,
                        "source": "page_scan"
                    })
            
            # Add all structured fees to the main fees list
            case_data["fees"].extend(structured_fees)
            
            # Extract any additional fee details for enhanced analysis
            # Look for daily rates
            daily_rate_match = re.search(r'(\$\d+(\.\d{2})?)\s*(?:per|a|each)\s*day', all_text, re.IGNORECASE)
            if daily_rate_match:
                case_data["dailyRate"] = daily_rate_match.group(1)
                logging.info(f"Found daily rate: {case_data['dailyRate']}")
            
            # Look for storage duration
            storage_days_match = re.search(r'(\d+)\s*days?\s*(?:of|for)?\s*storage', all_text, re.IGNORECASE)
            if storage_days_match:
                case_data["storageDays"] = int(storage_days_match.group(1))
                logging.info(f"Found storage days: {case_data['storageDays']}")
                
            # Look for invoice or reference numbers
            invoice_match = re.search(r'(?:invoice|reference|ref)(?:\s*#|number|num|:)\s*([A-Z0-9-]+)', all_text, re.IGNORECASE)
            if invoice_match:
                case_data["invoiceNumber"] = invoice_match.group(1)
                logging.info(f"Found invoice number: {case_data['invoiceNumber']}")
            
            # Final filtering for zero-amount fees (to ensure none slipped through)
            filtered_fees = []
            zero_fee_count = 0
            
            for fee in case_data["fees"]:
                if fee["amount"] > 0:
                    filtered_fees.append(fee)
                else:
                    zero_fee_count += 1
            
            # Update the case_data with filtered fees
            if zero_fee_count > 0:
                logging.info(f"Filtered out {zero_fee_count} fees with zero amounts in final filter")
                case_data["fees"] = filtered_fees
            
            # Sort fees by amount (largest first)
            case_data["fees"].sort(key=lambda x: x["amount"], reverse=True)
            
            # Enhanced duplicate removal with broader signature
            unique_fees = []
            seen_signatures = set()
            
            for fee in case_data["fees"]:
                # Create signatures with varying levels of specificity
                amount_str = f"{fee['amount']:.2f}"  # Normalize amount to avoid floating point issues
                
                # Primary signature with amount and truncated description
                primary_sig = f"{amount_str}_{fee['description'][:30]}"
                
                # Alternative signatures to catch duplicates with minor differences
                alt_sig1 = f"{amount_str}_{fee.get('category', '')}"  # Amount and category
                alt_sig2 = amount_str  # Just the amount (if it's a very specific amount like 247.53)
                
                if (primary_sig not in seen_signatures and 
                    # Only use alt_sig checks for very specific amounts (unlikely to be coincidental)
                    (not alt_sig1 in seen_signatures or int(fee['amount']) == fee['amount']) and
                    (not alt_sig2 in seen_signatures or int(fee['amount']) == fee['amount'] or fee['amount'] < 100)):
                    
                    seen_signatures.add(primary_sig)
                    seen_signatures.add(alt_sig1)
                    # Only add very specific amounts to seen
                    if int(fee['amount']) != fee['amount']:  # Not a round number
                        seen_signatures.add(alt_sig2)
                    
                    unique_fees.append(fee)
            
            # Replace with deduplicated list
            if len(unique_fees) < len(case_data["fees"]):
                logging.info(f"Removed {len(case_data['fees']) - len(unique_fees)} duplicate fees")
                case_data["fees"] = unique_fees
            
            logging.info(f"Found {len(case_data['fees'])} fees")
            
            # Always try to return whatever data we've managed to extract, even if it's partial
            # Wrap the entire Updates tab section in a try-except
            updates = []
            try:
                logging.info("Case information extraction complete, proceeding to Updates tab regardless of any missing data")
                # First take a screenshot of the case page before clicking Updates tab
                await page.screenshot(path=os.path.join('debug', f'before_updates_tab_{case_id}.png'))
                
                # Find Updates tab using multiple strategies
                logging.info("Looking for Updates tab")
                try:
                    # Try various selectors for the Updates tab with better logging
                    update_tab_selectors = [
                        "a:has-text('Updates')",
                        "a:has-text('History')",
                        "#updates-tab", 
                        "a[href*='updates']",
                        "a[href*='history']",
                        "ul.nav-tabs a:nth-child(2)",  # Often the second tab
                        ".nav-link:has-text('Updates')",
                        ".tab:has-text('Updates')"
                    ]
                    
                    tab_clicked = False
                    for selector in update_tab_selectors:
                        try:
                            update_link = await page.query_selector(selector)
                            if update_link:
                                await update_link.click()
                                tab_clicked = True
                                logging.info(f"Clicked Updates tab using selector: {selector}")
                                break
                        except Exception as e:
                            logging.debug(f"Failed to click Updates tab with selector {selector}: {str(e)}")
                            continue
                    
                    if not tab_clicked:
                        # List all available tabs for debugging
                        all_tabs = await page.query_selector_all(".nav-link, .nav-item a, .tab, li a")
                        logging.info(f"Found {len(all_tabs)} potential tabs to try")
                        
                        # Save all visible tabs to debug log
                        tab_texts = []
                        for i, tab in enumerate(all_tabs):
                            try:
                                tab_text = await tab.text_content()
                                tab_texts.append(f"Tab {i}: '{tab_text}'")
                                
                                # Try clicking tabs that look like they might be updates
                                if tab_text and any(keyword in tab_text.lower() for keyword in ['update', 'history', 'activity', 'log']):
                                    await tab.click()
                                    tab_clicked = True
                                    logging.info(f"Clicked tab with text: {tab_text}")
                                    break
                            except Exception as e:
                                logging.debug(f"Failed to process tab {i}: {str(e)}")
                        
                        logging.info(f"Available tabs: {', '.join(tab_texts)}")
                        
                        if not tab_clicked:
                            logging.warning("Could not find Updates tab, will attempt to find update data on current page")
                
                except Exception as e:
                    logging.warning(f"Error finding Updates tab: {str(e)}")
                
                # Wait for updates to load
                try:
                    await page.wait_for_load_state("networkidle", timeout=15000)  # Increased timeout
                    logging.info("Updates tab content loaded successfully")
                except Exception as e:
                    logging.warning(f"Timeout waiting for networkidle in Updates tab, continuing anyway: {str(e)}")
                    # Wait a reasonable time anyway
                    await page.wait_for_timeout(3000)
                await page.screenshot(path=os.path.join('debug', f'updates_tab_{case_id}.png'))
                
                # Function to extract updates from the current page
                async def extract_updates_from_page(page_num=1):
                    # Extract page content
                    updates_content = await page.content()
                    
                    # Save the content to a debug file for this page
                    with open(os.path.join('debug', f'updates_{case_id}_page{page_num}.html'), 'w', encoding='utf-8') as f:
                        f.write(updates_content)
                    
                    # Take screenshot of the current page
                    await page.screenshot(path=os.path.join('debug', f'updates_{case_id}_page{page_num}.png'))
                    
                    # Parse the page content
                    updates_soup = BeautifulSoup(updates_content, 'html.parser')
                    
                    # Find update elements using multiple strategies
                    update_elements = []
                    update_details_elements = []  # To store dt/dd pairs with Details
                    
                    # Strategy 0 (New): Look for dt/dd pairs where dt contains "Details"
                    # This specifically targets the structure provided in the user's example
                    dt_elements = updates_soup.find_all('dt', string=lambda s: s and 'Details' in s)
                    for dt in dt_elements:
                        # Look for the corresponding dd element which contains the details text
                        # First try to find a sibling dd
                        next_elem = dt
                        while next_elem and next_elem.name != 'dd':
                            next_elem = next_elem.next_sibling
                        
                        # If found a sibling dd, add it
                        if next_elem and next_elem.name == 'dd':
                            update_details_elements.append(next_elem)
                            logging.info(f"Found details dd element as sibling of dt")
                            continue
                            
                        # If not found as sibling, try to find via parent
                        if dt.parent:
                            dd_elements = dt.parent.find_all('dd')
                            if dd_elements:
                                update_details_elements.append(dd_elements[0])
                                logging.info(f"Found details dd element via parent")
                                continue
                                
                        # If still not found, try looking in nearby structure
                        if dt.parent and dt.parent.parent:
                            dd_elements = dt.parent.parent.find_all('dd')
                            if dd_elements:
                                update_details_elements.append(dd_elements[0])
                                logging.info(f"Found details dd element via grandparent")
                    
                    logging.info(f"Found {len(update_details_elements)} dd elements with details content")
                    
                    # Strategy 1: Look for elements with specific classes
                    class_patterns = [
                        r'update|history|log|activity',  # Common update-related classes
                        r'row|line|record|entry',        # Generic container classes that might hold updates
                        r'fee|transaction|payment'       # Fee-specific classes
                    ]
                    
                    for pattern in class_patterns:
                        elements = updates_soup.find_all(['div', 'tr', 'li'], class_=re.compile(pattern, re.IGNORECASE))
                        if elements:
                            update_elements.extend(elements)
                            logging.info(f"Found {len(elements)} elements with class pattern: {pattern}")
                    
                    # Strategy 2: Look for table rows if no elements found yet
                    if not update_elements and not update_details_elements:
                        logging.info("No update elements found with class patterns, trying tables")
                        tables = updates_soup.find_all('table')
                        for table in tables:
                            # Skip header row if it exists
                            rows = table.find_all('tr')[1:] if table.find('th') else table.find_all('tr')
                            if rows:
                                update_elements.extend(rows)
                                logging.info(f"Found {len(rows)} rows from table")
                    
                    # Strategy 3: Look for any elements containing date patterns and dollar amounts
                    if not update_elements and not update_details_elements:
                        logging.info("Still no update elements found, searching for elements with dates and amounts")
                        all_elements = updates_soup.find_all(['div', 'p', 'li', 'span'])
                        for element in all_elements:
                            text = element.get_text()
                            # Check if element has both a date and dollar amount
                            if re.search(r'\d{1,2}/\d{1,2}/\d{4}', text) and re.search(r'\$\d+(\.\d{2})?', text):
                                update_elements.append(element)
                        
                        if update_elements:
                            logging.info(f"Found {len(update_elements)} elements with dates and amounts")
                    
                    # Process all found elements
                    page_updates = []
                    
                    # Process the special dt/dd pairs with details first (from user example)
                    for dd_element in update_details_elements:
                        dd_text = dd_element.get_text().strip()
                        
                        # Skip empty elements
                        if not dd_text:
                            continue
                            
                        # Look for the parent container to get additional metadata like date, type, etc.
                        container = dd_element.find_parent(['dl', 'div', 'section'])
                        container_text = container.get_text().strip() if container else dd_text
                        
                        # Extract date using regex (support multiple formats)
                        date_match = re.search(r'\d{1,2}/\d{1,2}/\d{4}|\d{4}-\d{2}-\d{2}', container_text)
                        date = date_match.group(0) if date_match else "Unknown"
                        
                        # Extract update type (if available)
                        update_type = "Unknown"
                        update_type_dt = container.find('dt', string=lambda s: s and 'Update Type' in s) if container else None
                        if update_type_dt:
                            # Find corresponding dd
                            next_elem = update_type_dt
                            while next_elem and next_elem.name != 'dd':
                                next_elem = next_elem.next_sibling
                            if next_elem and next_elem.name == 'dd':
                                update_type = next_elem.get_text().strip()
                            else:
                                # Try parent method
                                if update_type_dt.parent:
                                    dd_elements = update_type_dt.parent.find_all('dd')
                                    if dd_elements:
                                        update_type = dd_elements[0].get_text().strip()
                        
                        # Extract fee amount using dollar pattern from the details text
                        # This is key for the user's example - extract dollar amount from the details
                        amount_match = re.search(dollar_pattern, dd_text)
                        amount_str = amount_match.group(0) if amount_match else "$0.00"
                        amount = float(amount_match.group(1).replace(',', '')) if amount_match else 0.0
                        
                        # Only include updates with dollar amounts
                        if amount > 0:
                            # Enhanced fee type identification with confidence score
                            fee_type_info = identify_fee_type(dd_text)
                            fee_type = fee_type_info["category"]
                            confidence = fee_type_info["confidence"]
                            color = fee_type_info["color"]
                            
                            # Enhanced fee status detection
                            fee_status = identify_fee_status(dd_text)
                            
                            logging.info(f"Adding update from dt/dd with amount: ${amount:.2f}")
                            
                            # Create update entry
                            update_entry = {
                                "date": date,
                                "details": dd_text,
                                "amount": amount,
                                "amountStr": amount_str,
                                "feeType": fee_type,
                                "feeTypeConfidence": confidence,
                                "feeTypeColor": color,
                                "status": fee_status,
                                "originalText": container_text,
                                "page": page_num,
                                "source": "details_dd"
                            }
                            
                            # Extract additional fee-specific data
                            # For storage fees
                            if "storage" in dd_text.lower():
                                # Look for number of days
                                days_match = re.search(r'(\d+)\s*days?', dd_text, re.IGNORECASE)
                                if days_match:
                                    update_entry["storageDays"] = int(days_match.group(1))
                                
                                # Look for daily rate
                                rate_match = re.search(r'(\$\d+(\.\d{2})?)\s*(?:per|a|each)\s*day', dd_text, re.IGNORECASE)
                                if rate_match:
                                    update_entry["dailyRate"] = rate_match.group(1)
                            
                            # For repossession fees
                            if any(term in dd_text.lower() for term in ["repo", "repossession", "recovery"]):
                                # Look for vehicle information
                                vehicle_match = re.search(r'(\d{4})\s+([A-Za-z]+)', dd_text)
                                if vehicle_match:
                                    update_entry["vehicleYear"] = vehicle_match.group(1)
                                    update_entry["vehicleMake"] = vehicle_match.group(2)
                            
                            page_updates.append(update_entry)
                    
                    # Process other update elements
                    for element in update_elements:
                        element_text = element.get_text().strip()
                        
                        # Skip empty elements or very short text
                        if not element_text or len(element_text) < 10:
                            continue
                            
                        # Skip header rows
                        if any(header.lower() in element_text.lower() for header in ['date', 'type', 'amount', 'detail']) and len(element_text) < 50:
                            continue
                        
                        # Extract date using regex (support multiple formats)
                        date_match = re.search(r'\d{1,2}/\d{1,2}/\d{4}|\d{4}-\d{2}-\d{2}', element_text)
                        date = date_match.group(0) if date_match else "Unknown"
                        
                        # Extract fee amount using dollar pattern
                        amount_match = re.search(dollar_pattern, element_text)
                        amount_str = amount_match.group(0) if amount_match else "$0.00"
                        amount = float(amount_match.group(1).replace(',', '')) if amount_match else 0.0
                        
                        # Only include updates with dollar amounts
                        if amount > 0:
                            # Get details (text excluding date and amount)
                            details = element_text
                            if date_match:
                                details = details.replace(date_match.group(0), "")
                            if amount_match:
                                details = details.replace(amount_match.group(0), "")
                            details = ' '.join(details.split())  # Normalize whitespace
                            
                            # Enhanced fee type identification with confidence score
                            fee_type_info = identify_fee_type(element_text)
                            fee_type = fee_type_info["category"]
                            confidence = fee_type_info["confidence"]
                            color = fee_type_info["color"]
                            
                            # Enhanced fee status detection
                            fee_status = identify_fee_status(element_text)
                            
                            logging.info(f"Adding update with amount: ${amount:.2f}")
                            
                            # Create a more comprehensive update object with detailed fee information
                            update_entry = {
                                "date": date,
                                "details": details.strip(),
                                "amount": amount,
                                "amountStr": amount_str,
                                "feeType": fee_type,
                                "feeTypeConfidence": confidence,
                                "feeTypeColor": color,
                                "status": fee_status,
                                "originalText": element_text,
                                "page": page_num,
                                "source": "general"
                            }
                            
                            # Extract additional fee-specific data
                            # For storage fees
                            if "storage" in element_text.lower():
                                # Look for number of days
                                days_match = re.search(r'(\d+)\s*days?', element_text, re.IGNORECASE)
                                if days_match:
                                    update_entry["storageDays"] = int(days_match.group(1))
                                
                                # Look for daily rate
                                rate_match = re.search(r'(\$\d+(\.\d{2})?)\s*(?:per|a|each)\s*day', element_text, re.IGNORECASE)
                                if rate_match:
                                    update_entry["dailyRate"] = rate_match.group(1)
                            
                            # For repossession fees
                            if any(term in element_text.lower() for term in ["repo", "repossession", "recovery"]):
                                # Look for vehicle information
                                vehicle_match = re.search(r'(\d{4})\s+([A-Za-z]+)', element_text)
                                if vehicle_match:
                                    update_entry["vehicleYear"] = vehicle_match.group(1)
                                    update_entry["vehicleMake"] = vehicle_match.group(2)
                            
                            page_updates.append(update_entry)
                    
                    return page_updates
                
                # Extract updates from the first page
                first_page_updates = await extract_updates_from_page(1)
                updates.extend(first_page_updates)
                logging.info(f"Extracted {len(first_page_updates)} updates from page 1")
                
                # Handle pagination - look for pagination controls and navigate through all pages
                pagination_found = False
                try:
                    # Check for various pagination elements
                    pagination_selectors = [
                        "ul.pagination", 
                        ".pagination", 
                        "nav[aria-label='pagination']", 
                        ".pager",
                        "div.pages",
                        ".page-numbers"
                    ]
                    
                    for selector in pagination_selectors:
                        pagination = await page.query_selector(selector)
                        if pagination:
                            pagination_found = True
                            logging.info(f"Found pagination with selector: {selector}")
                            
                            # Take a screenshot of the pagination area
                            await page.screenshot(path=os.path.join('debug', f'pagination_{case_id}.png'))
                            
                            # Try to find all page number links
                            page_links = await page.query_selector_all(f"{selector} a, {selector} button")
                            
                            if page_links:
                                logging.info(f"Found {len(page_links)} pagination links")
                                
                                # First approach: Try to click on numbered pages
                                page_num = 2  # Start from page 2 since we already processed page 1
                                max_pages = 20  # Safety limit to prevent infinite loops
                                
                                while page_num <= max_pages:
                                    # Try to find and click the link for the current page number
                                    next_page_clicked = False
                                    
                                    # Try multiple ways to find the next page link
                                    next_page_selectors = [
                                        f"{selector} a:has-text('{page_num}')",
                                        f"{selector} button:has-text('{page_num}')",
                                        f"{selector} li:nth-child({page_num+1}) a",  # +1 because first page is often 1
                                        f"a[data-page='{page_num}']",
                                        f"button[data-page='{page_num}']"
                                    ]
                                    
                                    for next_selector in next_page_selectors:
                                        try:
                                            next_link = await page.query_selector(next_selector)
                                            if next_link:
                                                # Click the link
                                                await next_link.click()
                                                logging.info(f"Clicked link to page {page_num}")
                                                
                                                # Wait for content to update
                                                try:
                                                    await page.wait_for_load_state("networkidle", timeout=10000)  # Increased timeout
                                                    logging.info(f"Pagination to page {page_num} successful")
                                                except Exception as e:
                                                    logging.warning(f"Timeout waiting for networkidle during pagination, continuing anyway: {str(e)}")
                                                    # Wait a reasonable time anyway
                                                    await page.wait_for_timeout(2000)
                                                
                                                # Extract updates from this page
                                                page_updates = await extract_updates_from_page(page_num)
                                                updates.extend(page_updates)
                                                logging.info(f"Extracted {len(page_updates)} updates from page {page_num}")
                                                
                                                next_page_clicked = True
                                                break
                                        except Exception as e:
                                            logging.debug(f"Failed to click {next_selector}: {str(e)}")
                                    
                                    # If we couldn't find a numbered link, try "Next" button
                                    if not next_page_clicked:
                                        next_btn_selectors = [
                                            "a:has-text('Next')",
                                            "button:has-text('Next')",
                                            ".pagination .next",
                                            "a[rel='next']",
                                            "a:has-text('»')"
                                        ]
                                        
                                        for next_btn in next_btn_selectors:
                                            try:
                                                btn = await page.query_selector(next_btn)
                                                if btn:
                                                    # Check if disabled
                                                    disabled = await btn.get_attribute("disabled")
                                                    aria_disabled = await btn.get_attribute("aria-disabled")
                                                    
                                                    if disabled == "true" or aria_disabled == "true":
                                                        logging.info("Next button is disabled, reached last page")
                                                        next_page_clicked = False
                                                        break
                                                    
                                                    # Click the Next button
                                                    await btn.click()
                                                    logging.info(f"Clicked 'Next' button for page {page_num}")
                                                    
                                                    # Wait for content to update
                                                    try:
                                                        await page.wait_for_load_state("networkidle", timeout=10000)  # Increased timeout
                                                        logging.info(f"Navigation via Next button to page {page_num} successful")
                                                    except Exception as e:
                                                        logging.warning(f"Timeout waiting for networkidle during Next button pagination, continuing anyway: {str(e)}")
                                                        # Wait a reasonable time anyway
                                                        await page.wait_for_timeout(2000)
                                                    
                                                    # Extract updates from this page
                                                    page_updates = await extract_updates_from_page(page_num)
                                                    updates.extend(page_updates)
                                                    logging.info(f"Extracted {len(page_updates)} updates from page {page_num} via Next button")
                                                    
                                                    next_page_clicked = True
                                                    break
                                            except Exception as e:
                                                logging.debug(f"Failed to click Next button {next_btn}: {str(e)}")
                                    
                                    # If we couldn't click any pagination element, we're done
                                    if not next_page_clicked:
                                        logging.info(f"No more pagination links available after page {page_num-1}")
                                        break
                                    
                                    # Move to next page
                                    page_num += 1
                            else:
                                logging.info("Found pagination container but no page links")
                            
                            break  # Exit the pagination selector loop
                    
                    if not pagination_found:
                        logging.info("No pagination controls found, all updates are on a single page")
                    
                except Exception as e:
                    logging.warning(f"Error handling pagination: {str(e)}")
                
                # Process and analyze the collected updates
                if updates:
                    # Remove any duplicate updates (based on date, amount, and details)
                    unique_updates = []
                    update_signatures = set()
                    
                    for update in updates:
                        # Normalize amount to avoid floating point comparison issues
                        amount_str = f"{update['amount']:.2f}"
                        
                        # Create multiple signatures for this update to catch different kinds of duplicates
                        # Primary signature with date, amount and details
                        primary_sig = f"{update['date']}_{amount_str}_{update['details'][:50]}"
                        
                        # Secondary signature with just date and amount (catches cases where details vary slightly)
                        secondary_sig = f"{update['date']}_{amount_str}"
                        
                        # Choose which signature to use based on specificity of the amount
                        # For very specific amounts, we can use just date+amount as a signature
                        # For common amounts (like $0, $100, etc.), we need more info to avoid false duplicates
                        is_specific_amount = (int(update['amount']) != update['amount'] or 
                                             update['amount'] > 1000 or 
                                             (update['amount'] > 0 and update['amount'] < 10))
                        
                        if primary_sig not in update_signatures:
                            # For specific amounts, also check the secondary signature
                            if not is_specific_amount or secondary_sig not in update_signatures:
                                update_signatures.add(primary_sig)
                                update_signatures.add(secondary_sig)
                                unique_updates.append(update)
                    
                    if len(unique_updates) < len(updates):
                        logging.info(f"Removed {len(updates) - len(unique_updates)} duplicate updates")
                        updates = unique_updates
                    
                    # Sort updates by date (newest first)
                    try:
                        updates.sort(key=lambda x: datetime.datetime.strptime(x['date'], '%m/%d/%Y') 
                                     if x['date'] != "Unknown" else datetime.datetime(1900, 1, 1), 
                                     reverse=True)
                    except Exception as e:
                        logging.warning(f"Failed to sort updates by date: {str(e)}")
                    
                    # Save the full set of updates for debugging
                    with open(os.path.join('debug', f'all_updates_{case_id}.json'), 'w', encoding='utf-8') as f:
                        json.dump(updates, f, indent=2)
                
                logging.info(f"Finished extracting updates. Total count: {len(updates)}")
                
            except Exception as e:
                logging.exception(f"Error extracting updates: {str(e)}")
            
            # Return with successfully collected data, even if some portions failed
            try:
                await browser.close()
            except Exception as browser_close_error:
                logging.error(f"Error closing browser: {str(browser_close_error)}")
            
            # Log updates status
            logging.info(f"Final updates count for return: {len(updates)}")
            if len(updates) == 0:
                logging.warning("No updates found. This may be due to zero-filtering or extraction issues.")
            
            # Check for minimum viable data before considering it a success
            if case_data.get("clientName") != "Error extracting data" or case_data.get("lienHolder") != "Error extracting data":
                # Make sure we always have at least the updates structure
                if not updates:
                    logging.info("Creating empty updates structure to ensure UI gets updates data")
                    updates = []  # Ensure it's an empty list at minimum
                
                return True, {"case_data": case_data, "updates": updates}
            else:
                # We have no useful data
                return False, "Failed to extract any useful case data. Please try again."
            
        except Exception as e:
            logging.exception(f"Error extracting case data: {str(e)}")
            
            # Always try to close the browser
            try:
                await browser.close()
            except Exception:
                pass
                
            # Return partial data if we have any
            if 'case_data' in locals() and isinstance(case_data, dict) and len(case_data.get("fees", [])) > 0:
                logging.warning("Returning partial data despite error")
                return True, {"case_data": case_data, "updates": updates if 'updates' in locals() else []}
            else:
                return False, f"Error extracting case data: {str(e)}"

def lookup_repo_fee(client_name, lienholder_name, fee_type_name):
    """
    Lookup repo fee from database based on client name, lienholder name, and fee type.
    Based on the implementation from server-upgradedv2.py.

    Args:
        client_name (str): The name of the client
        lienholder_name (str): The name of the lienholder
        fee_type_name (str): The name of the fee type (e.g., 'Involuntary Repo')

    Returns:
        dict: A dictionary containing fee details, or None if no matching fee is found
    """
    logging.info(f'Looking up repo fee for: Client="{client_name}", Lienholder="{lienholder_name}", FeeType="{fee_type_name}"')

    # Make sure we're using the correct variable names
    case_client_name = client_name
    case_lienholder_name = lienholder_name
    case_repo_type = fee_type_name
    
    # Get database config
    try:
        db_config = app_config['database']
        
        # Build connection string from config
        server = db_config['server']
        database = db_config['database']
        username = db_config['username']
        password = db_config['password']
        
        # Build a proper pyodbc connection string
        conn_str = f"DRIVER={{SQL Server}};SERVER={server};DATABASE={database};UID={username};PWD={password}"
        logging.info(f"Connecting to database: {server}/{database}")
        
        # Connect to the database
        conn = pyodbc.connect(conn_str)
    except Exception as e:
        logging.error(f"Could not connect to database: {str(e)}")
        return None

    try:
        cursor = conn.cursor()

        # Step 1: Get foreign keys from names
        logging.info("Getting foreign keys from names...")
        cursor.execute("SELECT TOP 1 id FROM dbo.RDN_Client WHERE client_name = ?", case_client_name)
        client_row = cursor.fetchone()
        if not client_row:
            logging.warning(f"Client '{case_client_name}' not found in database")
            return None
        client_id = client_row[0]

        # Get lienholder ID
        cursor.execute("SELECT TOP 1 id FROM dbo.Lienholder WHERE lienholder_name = ?", case_lienholder_name)
        lienholder_row = cursor.fetchone()
        if not lienholder_row:
            logging.warning(f"Lienholder '{case_lienholder_name}' not found in database")
            lienholder_id = None  # We'll handle this in the fallback logic
        else:
            lienholder_id = lienholder_row[0]

        # Get fee type ID
        cursor.execute("SELECT TOP 1 id FROM dbo.FeeType WHERE fee_type_name = ?", case_repo_type)
        fee_type_row = cursor.fetchone()
        if not fee_type_row:
            logging.warning(f"Fee type '{case_repo_type}' not found in database")
            return None
        fee_type_id = fee_type_row[0]

        # Step 2: Check if a matching record exists and return it if found (primary logic)
        if lienholder_id:
            # Using parameterized query for safety
            query = """
                SELECT
                    fd.fd_id,
                    c.client_name,
                    lh.lienholder_name,
                    ft.fee_type_name,
                    fd.amount
                FROM dbo.FeeDetails2 fd
                JOIN dbo.RDN_Client c ON fd.client_id = c.id
                JOIN dbo.Lienholder lh ON fd.lh_id = lh.id
                JOIN dbo.FeeType ft ON fd.ft_id = ft.id
                WHERE fd.client_id = ? AND fd.lh_id = ? AND fd.ft_id = ?
            """
            logging.info(f"Executing primary lookup query with params: [{client_id}, {lienholder_id}, {fee_type_id}]")
            cursor.execute(query, [client_id, lienholder_id, fee_type_id])
            row = cursor.fetchone()

            if row:
                logging.info(f"Found matching fee record for specific lienholder '{lienholder_name}'")
                return {
                    'fd_id': row[0],
                    'client_name': row[1],
                    'lienholder_name': row[2],
                    'fee_type': row[3],
                    'amount': row[4],
                    'is_fallback': False
                }

        # Step 3: If no record found, try with 'Standard' lienholder (fallback logic)
        logging.info(f"No specific record found. Looking up 'Standard' lienholder as fallback...")
        cursor.execute("SELECT TOP 1 id FROM dbo.Lienholder WHERE lienholder_name = 'Standard'")
        standard_row = cursor.fetchone()

        if not standard_row:
            logging.error("'Standard' lienholder not found in database")
            return None

        standard_lienholder_id = standard_row[0]

        # Look up using Standard lienholder
        fallback_query = """
            SELECT
                fd.fd_id,
                c.client_name,
                lh.lienholder_name,
                ft.fee_type_name,
                fd.amount
            FROM dbo.FeeDetails2 fd
            JOIN dbo.RDN_Client c ON fd.client_id = c.id
            JOIN dbo.Lienholder lh ON fd.lh_id = lh.id
            JOIN dbo.FeeType ft ON fd.ft_id = ft.id
            WHERE fd.client_id = ? AND fd.lh_id = ? AND fd.ft_id = ?
        """
        logging.info(f"Executing fallback lookup query with params: [{client_id}, {standard_lienholder_id}, {fee_type_id}]")
        cursor.execute(fallback_query, [client_id, standard_lienholder_id, fee_type_id])
        fallback_row = cursor.fetchone()

        if fallback_row:
            logging.info(f"Found fallback fee using 'Standard' lienholder")
            return {
                'fd_id': fallback_row[0],
                'client_name': fallback_row[1],
                'lienholder_name': fallback_row[2] + " (Standard Fallback)",
                'fee_type': fallback_row[3],
                'amount': fallback_row[4],
                'is_fallback': True,
                'message': f"Lienholder '{case_lienholder_name}' specific fee not found. Using Standard amount."
            }

        logging.warning("No fee record found with either specific lienholder or fallback")
        return None

    except Exception as e:
        logging.error(f"Error looking up repo fee: {str(e)}")
        import traceback
        logging.error(traceback.format_exc())
        return None

    finally:
        if conn:
            conn.close()

@app.route('/api/query-database', methods=['GET'])
def query_database():
    """Query Azure SQL database for fee information using config credentials"""
    logging.info("Database query request received")
    if 'case_data' not in session:
        logging.error("Case data not available")
        return jsonify({"success": False, "message": "Case data not available"})
    
    case_data = session.get('case_data')
    
    try:
        # Get client information from case data
        client_name = case_data.get('clientName')
        lienholder_name = case_data.get('lienHolder')
        
        # For repo type, either use from case data or default to 'Involuntary Repo'
        repo_type = case_data.get('repoType', 'Involuntary Repo')
        if not repo_type or repo_type == 'Not Found':
            repo_type = 'Involuntary Repo'
        
        logging.info(f"Looking up database fee for: Client={client_name}, Lienholder={lienholder_name}, FeeType={repo_type}")
        
        # Use the enhanced lookup_repo_fee function
        db_result = lookup_repo_fee(client_name, lienholder_name, repo_type)
        
        # If no result, create a mock result for testing
        if not db_result:
            logging.warning("No database result found, using mock data")
            db_result = {
                "fd_id": f"FD-{case_data.get('caseId', '0000')}",
                "client_name": client_name,
                "lienholder_name": lienholder_name,
                "fee_type": repo_type,
                "amount": 350.00,
                "is_fallback": False
            }
        
        # Format the result to match the expected format for the client
        api_result = {
            "fdId": db_result['fd_id'],
            "clientName": db_result['client_name'],
            "lienholderName": db_result['lienholder_name'],
            "feeTypeName": db_result['fee_type'],
            "amount": float(db_result['amount']) if isinstance(db_result['amount'], decimal.Decimal) else db_result['amount'],
            "isFallback": db_result.get('is_fallback', False)
        }
        
        # If there's a message, include it
        if 'message' in db_result:
            api_result["message"] = db_result['message']
            
        logging.info(f"Database result: {api_result}")
        
        # Store in session
        session['db_data'] = api_result
        
        return jsonify({"success": True, "data": api_result})
        
    except Exception as e:
        logging.exception(f"Database error: {str(e)}")
        # Create mock data for testing when database connection fails
        mock_data = {
            "fdId": f"FD-{case_data.get('caseId', '0000')}",
            "clientName": case_data.get('clientName'),
            "lienholderName": case_data.get('lienHolder'),
            "feeTypeName": "Involuntary Repo",
            "amount": 350.00,
            "isFallback": True,
            "message": "Database connection failed. Using mock data."
        }
        
        session['db_data'] = mock_data
        
        return jsonify({
            "success": True, 
            "data": mock_data, 
            "warning": f"Database connection failed: {str(e)}. Using mock data."
        })

@app.route('/api/updates', methods=['GET'])
def get_updates():
    """Retrieve updates data from session"""
    logging.info("Updates data request received")
    if 'updates' not in session:
        logging.error("Updates data not available")
        return jsonify({"success": False, "message": "Updates data not available"})
    
    # Return all updates - filtering will be handled in the frontend
    # This ensures we don't lose any data that might be needed
    return jsonify({"success": True, "data": session.get('updates')})

def auto_fetch_database_fees(case_data=None):
    """
    Automatically fetch fee information from database using current case info
    This function will be called after case information is extracted
    
    Args:
        case_data (dict, optional): The case data containing client info. If None, uses session data.
        
    Returns:
        dict: The database fee data, or None if not found
    """
    # Get case data from session if not provided
    if case_data is None:
        if 'case_data' not in session:
            logging.error("Case data not available for auto database fetch")
            return None
        case_data = session.get('case_data')
    
    client_name = case_data.get('clientName')
    lienholder_name = case_data.get('lienHolder')
    repo_type = case_data.get('repoType', 'Involuntary Repo')
    
    # Default to Involuntary Repo if not found or invalid
    if not repo_type or repo_type == 'Not Found':
        repo_type = 'Involuntary Repo'
    
    logging.info(f"Auto-fetching fee data from database using case information:")
    logging.info(f"Client: {client_name}")
    logging.info(f"Lienholder: {lienholder_name}")
    logging.info(f"Repo Type: {repo_type}")
    
    # Fetch matching fee records from database
    try:
        db_result = lookup_repo_fee(client_name, lienholder_name, repo_type)
        
        if db_result:
            logging.info(f"Successfully fetched repo fee from database: ${float(db_result['amount']):.2f}")
            
            # Format for frontend API
            api_result = {
                "fdId": db_result['fd_id'],
                "clientName": db_result['client_name'],
                "lienholderName": db_result['lienholder_name'],
                "feeTypeName": db_result['fee_type'],
                "amount": float(db_result['amount']) if isinstance(db_result['amount'], decimal.Decimal) else db_result['amount'],
                "isFallback": db_result.get('is_fallback', False)
            }
            
            # If there's a message, include it
            if 'message' in db_result:
                api_result["message"] = db_result['message']
                
            # Store in session
            session['db_data'] = api_result
            
            return api_result
        else:
            logging.warning("No matching fee data found in database")
            
            # Create mock data when no database match
            mock_data = {
                "fdId": f"FD-{case_data.get('caseId', '0000')}",
                "clientName": client_name,
                "lienholderName": lienholder_name,
                "feeTypeName": repo_type,
                "amount": 350.00,
                "isFallback": True,
                "message": "No matching database record found. Using default amount."
            }
            
            # Store in session
            session['db_data'] = mock_data
            
            return mock_data
            
    except Exception as e:
        logging.error(f"Error in auto database fetch: {str(e)}")
        logging.error(traceback.format_exc())
        
        # Create mock data for exception cases
        mock_data = {
            "fdId": f"FD-{case_data.get('caseId', '0000')}",
            "clientName": client_name,
            "lienholderName": lienholder_name,
            "feeTypeName": repo_type,
            "amount": 350.00,
            "isFallback": True,
            "message": f"Database error: {str(e)}. Using default amount."
        }
        
        # Store in session
        session['db_data'] = mock_data
        
        return mock_data

@app.route('/api/results', methods=['GET'])
def get_results():
    """Retrieve all results data"""
    logging.info("Results request received")
    if 'case_data' not in session or 'updates' not in session:
        logging.error("Complete data not available")
        return jsonify({"success": False, "message": "Complete data not available"})
    
    # Ensure we have database data - if not, try to auto-fetch it
    if 'db_data' not in session:
        logging.info("Database data not in session, attempting auto-fetch")
        db_data = auto_fetch_database_fees()
    else:
        db_data = session.get('db_data')
    
    # Get case data and filter out fees with zero amounts
    case_data = session.get('case_data', {})
    if "fees" in case_data:
        case_data = case_data.copy()  # Create a copy to avoid modifying the session directly
        case_data["fees"] = [fee for fee in case_data["fees"] if fee.get('amount', 0) > 0]
        logging.info(f"Filtered case fees in results to exclude zero amounts. Remaining fees: {len(case_data['fees'])}")
    
    # Return all updates without filtering to ensure nothing is missing
    updates = session.get('updates', [])
    
    return jsonify({
        "success": True,
        "caseData": case_data,
        "dbData": db_data,
        "updates": updates
    })

@app.route('/api/export/excel', methods=['GET'])
def export_excel():
    """Generate Excel export of results"""
    logging.info("Excel export request received")
    if 'case_data' not in session or 'db_data' not in session or 'updates' not in session:
        logging.error("Complete data not available for export")
        return jsonify({"success": False, "message": "Complete data not available"})
    
    # Get case data and filter out fees with zero amounts
    case_data = session.get('case_data').copy()  # Create a copy to avoid modifying the session
    if "fees" in case_data:
        case_data["fees"] = [fee for fee in case_data["fees"] if fee.get('amount', 0) > 0]
        logging.info(f"Filtered case fees in export to exclude zero amounts. Remaining fees: {len(case_data['fees'])}")
    
    # Keep all updates to maintain consistency with the updates tab
    updates = session.get('updates', [])
    
    db_data = session.get('db_data')
    
    try:
        # Create a file name
        file_name = f"JamiBilling_Case_{case_data['caseId']}_{datetime.datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        file_path = os.path.join('static', 'exports', file_name)
        logging.info(f"Creating Excel export: {file_path}")
        
        # Create Excel workbook
        workbook = openpyxl.Workbook()
        
        # Create summary sheet
        summary_sheet = workbook.active
        summary_sheet.title = "Summary"
        
        # Add headers
        summary_sheet['A1'] = "Item"
        summary_sheet['B1'] = "Value"
        
        # Add data
        summary_data = [
            ["Case ID", case_data['caseId']],
            ["Client Name", case_data['clientName']],
            ["Lien Holder", case_data['lienHolder']],
            ["Order To", case_data['orderTo']],
            ["", ""],
            ["Database Information", ""],
            ["Fee ID", db_data['fdId']],
            ["Fee Type", db_data['feeTypeName']],
            ["Amount", db_data['amount']],
            ["", ""],
            ["Total Fees", sum(fee['amount'] for fee in case_data['fees'])]
        ]
        
        # Add rows to sheet
        for i, row in enumerate(summary_data, 2):  # Start from row 2
            summary_sheet[f'A{i}'] = row[0]
            summary_sheet[f'B{i}'] = row[1]
        
        # Create enhanced fees sheet with all fee details
        fees_sheet = workbook.create_sheet(title="Fees")
        
        # Add enhanced headers for more comprehensive fee data
        fees_sheet['A1'] = "Description"
        fees_sheet['B1'] = "Category"
        fees_sheet['C1'] = "Amount"
        fees_sheet['D1'] = "Status"
        fees_sheet['E1'] = "Source"
        fees_sheet['F1'] = "Confidence"
        fees_sheet['G1'] = "Notes"
        
        # Add enhanced fee data with all available fields
        for i, fee in enumerate(case_data['fees'], 2):  # Start from row 2
            fees_sheet[f'A{i}'] = fee.get('description', '')
            fees_sheet[f'B{i}'] = fee.get('category', '')
            fees_sheet[f'C{i}'] = fee.get('amount', 0)
            fees_sheet[f'D{i}'] = fee.get('status', '')
            fees_sheet[f'E{i}'] = fee.get('source', '')
            fees_sheet[f'F{i}'] = fee.get('confidence', '')
            
            # Compile any additional fee-specific data into notes
            notes = []
            if 'dailyRate' in fee:
                notes.append(f"Daily Rate: {fee['dailyRate']}")
            if 'storageDays' in fee:
                notes.append(f"Storage Days: {fee['storageDays']}")
            if 'vehicleYear' in fee and 'vehicleMake' in fee:
                notes.append(f"Vehicle: {fee['vehicleYear']} {fee['vehicleMake']}")
                
            fees_sheet[f'G{i}'] = "; ".join(notes)
        
        # Apply formatting to make the sheet more readable
        # Auto-adjust column widths to fit content
        for col in ['A', 'B', 'D', 'E', 'G']:
            max_length = 0
            for cell in fees_sheet[col]:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)  # Cap at 50 characters
            fees_sheet.column_dimensions[col].width = adjusted_width
        
        # Create an enhanced updates sheet with all available fields
        updates_sheet = workbook.create_sheet(title="Updates")
        
        # Add enhanced headers
        updates_sheet['A1'] = "Date"
        updates_sheet['B1'] = "Details"
        updates_sheet['C1'] = "Fee Type"
        updates_sheet['D1'] = "Amount"
        updates_sheet['E1'] = "Status"
        updates_sheet['F1'] = "Confidence"
        updates_sheet['G1'] = "Page"
        updates_sheet['H1'] = "Additional Info"
        
        # Add enhanced update data with all available fields
        for i, update in enumerate(updates, 2):  # Start from row 2
            updates_sheet[f'A{i}'] = update.get('date', '')
            updates_sheet[f'B{i}'] = update.get('details', '')
            updates_sheet[f'C{i}'] = update.get('feeType', '')
            updates_sheet[f'D{i}'] = update.get('amount', 0)
            updates_sheet[f'E{i}'] = update.get('status', '')
            updates_sheet[f'F{i}'] = update.get('feeTypeConfidence', '')
            updates_sheet[f'G{i}'] = update.get('page', '')
            
            # Compile any additional data fields into notes
            notes = []
            if 'dailyRate' in update:
                notes.append(f"Daily Rate: {update['dailyRate']}")
            if 'storageDays' in update:
                notes.append(f"Storage Days: {update['storageDays']}")
            if 'vehicleYear' in update and 'vehicleMake' in update:
                notes.append(f"Vehicle: {update['vehicleYear']} {update['vehicleMake']}")
                
            updates_sheet[f'H{i}'] = "; ".join(notes)
        
        # Apply formatting to make the sheet more readable
        # Auto-adjust column widths
        for col in ['A', 'B', 'C', 'E', 'H']:
            max_length = 0
            for cell in updates_sheet[col]:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)  # Cap at 50 characters
            updates_sheet.column_dimensions[col].width = adjusted_width
            
        # Create a fee summary sheet to provide totals by fee category
        summary_fees_sheet = workbook.create_sheet(title="Fee Summary")
        
        # Add headers
        summary_fees_sheet['A1'] = "Fee Category"
        summary_fees_sheet['B1'] = "Total Amount"
        summary_fees_sheet['C1'] = "Count"
        summary_fees_sheet['D1'] = "Details"
        
        # Group fees by category and calculate totals
        fee_categories = {}
        for fee in case_data['fees']:
            category = fee.get('category', 'Unknown')
            if category not in fee_categories:
                fee_categories[category] = {
                    'total': 0,
                    'count': 0,
                    'details': []
                }
            
            fee_categories[category]['total'] += fee.get('amount', 0)
            fee_categories[category]['count'] += 1
            fee_categories[category]['details'].append(
                f"{fee.get('description','')[:50]}... ({fee.get('amountStr', '')})"
            )
        
        # Add fee category summary data
        for i, (category, data) in enumerate(fee_categories.items(), 2):
            summary_fees_sheet[f'A{i}'] = category
            summary_fees_sheet[f'B{i}'] = data['total']
            summary_fees_sheet[f'C{i}'] = data['count']
            summary_fees_sheet[f'D{i}'] = "; ".join(data['details'][:3])  # Show top 3 fees
            
        # Format the summary sheet
        for col in ['A', 'D']:
            summary_fees_sheet.column_dimensions[col].width = 40
        
        # Save the Excel file
        workbook.save(file_path)
        logging.info("Excel file saved successfully")
        
        # Return the file URL
        file_url = f"/static/exports/{file_name}"
        return jsonify({"success": True, "file_url": file_url})
        
    except Exception as e:
        logging.exception(f"Error generating Excel: {str(e)}")
        return jsonify({"success": False, "message": f"Error generating Excel: {str(e)}"})

def identify_fee_type(text):
    """Identify fee type from text description"""
    if not text:
        return {"category": "Unknown", "confidence": 0, "color": "#858796"}
    
    text_lower = text.lower()
    highest_confidence = 0
    detected_category = "Unknown"
    category_color = "#858796"  # Default gray
    
    # Check each category's keywords against the text
    for category, data in fee_categories.items():
        for keyword in data["keywords"]:
            if keyword.lower() in text_lower:
                # Found an exact match
                return {
                    "category": category,
                    "confidence": 1,
                    "color": data["color"]
                }
        
        # No exact match, try partial matching with score
        score = 0
        for keyword in data["keywords"]:
            keyword_parts = keyword.lower().split()
            for part in keyword_parts:
                if len(part) > 3 and part in text_lower:
                    score += 0.5
        
        if score > highest_confidence:
            highest_confidence = score
            detected_category = category
            category_color = data["color"]
    
    if highest_confidence > 0:
        return {
            "category": detected_category,
            "confidence": min(highest_confidence, 0.9),  # Cap at 0.9 for partial matches
            "color": category_color
        }
    
    return {"category": "Unknown", "confidence": 0, "color": "#858796"}

def identify_fee_status(text):
    """Identify fee status from text description"""
    if not text:
        return "Unknown"
    
    text_lower = text.lower()
    
    status_keywords = {
        "Paid": ["paid", "payment received", "payment complete"],
        "Not Paid": ["not paid", "unpaid", "payment pending"],
        "Approved": ["approved", "accepted", "authorized"],
        "Pending": ["pending", "awaiting", "in process"],
        "Denied": ["denied", "rejected", "declined"],
        "Waived": ["waived", "forgiven", "no charge"]
    }
    
    for status, keywords in status_keywords.items():
        for keyword in keywords:
            if keyword in text_lower:
                return status
    
    return "Unknown"

@app.route('/debug-logs', methods=['GET'])
def view_debug_logs():
    """View debug information (only in development)"""
    if not app.debug:
        return jsonify({"error": "Debug mode not enabled"}), 403
    
    try:
        # Get list of debug files
        debug_files = []
        for filename in os.listdir('debug'):
            if filename.endswith('.png') or filename.endswith('.html'):
                file_path = os.path.join('debug', filename)
                file_info = {
                    'name': filename,
                    'path': file_path,
                    'size': os.path.getsize(file_path),
                    'modified': datetime.datetime.fromtimestamp(os.path.getmtime(file_path)).strftime('%Y-%m-%d %H:%M:%S')
                }
                debug_files.append(file_info)
        
        # Sort by modified date (newest first)
        debug_files.sort(key=lambda x: x['modified'], reverse=True)
        
        # Get log file content
        log_content = ""
        if os.path.exists('jami_billing.log'):
            with open('jami_billing.log', 'r') as f:
                log_content = f.read()
        
        return jsonify({
            "debug_files": debug_files,
            "log_content": log_content
        })
    except Exception as e:
        return jsonify({"error": str(e)}), 500

@app.route('/debug-file/<filename>', methods=['GET'])
def get_debug_file(filename):
    """Get a specific debug file (only in development)"""
    if not app.debug:
        return jsonify({"error": "Debug mode not enabled"}), 403
    
    try:
        # Ensure filename is safe and exists
        safe_filename = secure_filename(filename)
        file_path = os.path.join('debug', safe_filename)
        
        if not os.path.exists(file_path):
            return jsonify({"error": "File not found"}), 404
        
        # Serve the file
        if filename.endswith('.png'):
            return send_file(file_path, mimetype='image/png')
        elif filename.endswith('.html'):
            return send_file(file_path, mimetype='text/html')
        else:
            return send_file(file_path, mimetype='text/plain')
    except Exception as e:
        return jsonify({"error": str(e)}), 500

@app.route('/healthcheck', methods=['GET'])
def healthcheck():
    """Simple health check endpoint"""
    return jsonify({
        "status": "ok",
        "version": "1.0",
        "timestamp": datetime.datetime.now().isoformat()
    })

if __name__ == '__main__':
    # Create necessary directories
    os.makedirs('debug', exist_ok=True)
    os.makedirs('static/exports', exist_ok=True)
    os.makedirs('flask_session', exist_ok=True)
    
    # Initialize event loop
    get_event_loop()
    
    # Log startup
    logging.info("Starting JamiBilling application")
    
    app.run(debug=True, port=5000)